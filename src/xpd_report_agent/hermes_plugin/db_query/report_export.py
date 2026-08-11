from __future__ import annotations

import csv
import fcntl
import io
import json
import math
import os
import re
import shutil
import threading
import time
import unicodedata
import uuid
import zipfile
from contextlib import contextmanager
from copy import copy
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from xml.sax.saxutils import escape
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from .query_results import query_result_registry
from .report_oss import upload_report_artifact

ARTIFACT_ID_PATTERN = re.compile(r"art_[0-9a-f]{32}")
ARTIFACT_FILENAME_PATTERN = re.compile(r"art_[0-9a-f]{32}__[^/\\]+\.(?:csv|xlsx|md|pdf|json)", re.I)
SESSION_ID_PATTERN = re.compile(r"xpd_[0-9a-f]{20}_[A-Za-z0-9_]+")
FILENAME_SEPARATOR = "__"
SUPPORTED_FORMATS = frozenset({"csv", "xlsx", "markdown", "pdf", "json"})
ANALYSIS_TYPES = frozenset({"simple", "comparison", "diagnostic"})
FORMAT_EXTENSIONS = {
    "csv": ".csv",
    "xlsx": ".xlsx",
    "markdown": ".md",
    "pdf": ".pdf",
    "json": ".json",
}
MEDIA_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "markdown": "text/markdown; charset=utf-8",
    "pdf": "application/pdf",
    "json": "application/json; charset=utf-8",
}
MAX_TEXT_CHARS = 20_000
MAX_LIST_ITEMS = 30
MAX_TREND_ITEMS = 100
MAX_COLUMNS = 256
MAX_FILENAME_BYTES = 200
EXCEL_MAX_CELL_CHARS = 32_767
EXCEL_MAX_EXACT_INTEGER = 999_999_999_999_999
EXCEL_ILLEGAL_CONTROL_PATTERN = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")

CHINESE_COLUMN_LABELS = {
    "caliber": "统计口径",
    "grain": "统计粒度",
    "period_label": "统计时间",
    "period_type": "对比周期",
    "comparison_period": "对比周期",
    "start_date": "直播开始日期",
    "start_time": "开始时间",
    "end_time": "结束时间",
    "event_time": "发生时间",
    "session_cnt": "直播场次数",
    "live_cnt": "直播场次数",
    "duration_min": "直播时长（分钟）",
    "total_hours": "总直播时长（小时）",
    "item_id": "商品ID",
    "stat_date": "统计日期",
    "item_title": "商品标题",
    "item_title_hash": "商品标题哈希",
    "live_session_id": "直播场次ID",
    "session_title": "直播标题",
    "live_title": "直播标题",
    "live_start_time": "直播开始时间",
    "live_end_time": "直播结束时间",
    "live_duration_min": "直播时长（分钟）",
    "item_exposure_pv": "商品曝光次数",
    "item_exposure_uv": "商品曝光人数",
    "item_click_pv": "商品点击次数",
    "item_click_uv": "商品点击人数",
    "item_click_rate": "商品点击率",
    "channel_exposure_uv": "频道曝光人数",
    "channel_exposure_pv": "频道曝光次数",
    "channel_click_uv": "频道点击人数",
    "channel_click_pv": "频道点击次数",
    "cover_click_rate": "封面点击率",
    "is_digital_human_live": "是否数字人直播",
    "watch_uv": "观看人数",
    "watch_pv": "观看次数",
    "watch_total_duration": "总观看时长",
    "traffic_coupon_cost": "流量券成本",
    "avg_watch_sec_per_user": "人均观看时长（秒）",
    "avg_watch_sec_per_view": "次均观看时长（秒）",
    "peak_online_user_cnt": "最高在线人数",
    "avg_peak_online": "平均峰值在线人数",
    "new_fans_cnt": "新增粉丝数",
    "fan_conversion_rate": "转粉率",
    "interact_uv": "互动人数",
    "interact_cnt": "互动次数",
    "like_uv": "点赞人数",
    "like_cnt": "点赞次数",
    "comment_uv": "评论人数",
    "comment_cnt": "评论次数",
    "share_uv": "分享人数",
    "share_cnt": "分享次数",
    "cart_amt": "加购金额",
    "cart_byr_cnt": "加购买家数",
    "cart_cnt": "加购次数",
    "cart_itm_qty": "加购商品件数",
    "cart_conversion_rate": "加购转化率",
    "pay_amt": "成交金额",
    "gmv": "成交金额（GMV）",
    "pay_byr_cnt": "支付买家数",
    "pay_itm_qty": "支付商品件数",
    "pay_qty": "成交商品件数",
    "pay_conversion_rate": "支付转化率",
    "pay_ord_cnt": "支付订单数",
    "customer_unit_price": "客单价",
    "avg_customer_price": "平均客单价",
    "order_unit_price": "订单均价",
    "item_unit_price": "商品均价",
    "refund_amt": "退款金额",
    "refund_byr_cnt": "退款买家数",
    "refund_itm_qty": "退款商品件数",
    "refund_qty": "退款商品件数",
    "refund_ord_cnt": "退款订单数",
    "refund_byr_rate": "买家退款率",
    "return_goods_rate": "退货率",
    "refund_order_rate": "订单退款率",
    "refund_rate": "退货率",
    "refund_rate_pct": "金额退款率",
    "refund_qty_rate": "件数退货率",
    "refund_amt_rate": "金额退货率",
    "qty_refund_rate_pct": "件数退货率",
    "amt_refund_rate_pct": "金额退货率",
    "item_click_rate_pct": "商品点击率",
    "pay_conv_rate_pct": "支付转化率",
    "confirm_amt": "确认收货金额",
    "confirm_byr_cnt": "确认收货买家数",
    "confirm_ord_cnt": "确认收货订单数",
    "confirm_itm_qty": "确认收货商品件数",
    "subpay_amt": "预售定金金额",
    "subpay_byr_cnt": "预售定金买家数",
    "subpay_ord_cnt": "预售定金订单数",
    "subpay_itm_qty": "预售定金商品件数",
    "predict_amt": "预计成交金额",
    "tail_pay_amt": "预售尾款金额",
    "tail_pay_ord_cnt": "预售尾款订单数",
    "tail_predict_amt": "预计尾款金额",
    "subpay_tail_pay_amt": "定金及尾款金额",
    "subpay_tail_pay_ord_cnt": "定金及尾款订单数",
    "subpay_tail_predict_amt": "定金及预计尾款金额",
    "ord_cnt": "下单订单数",
    "ord_pay_amt": "下单口径成交金额",
    "ord_pay_ord_cnt": "下单口径支付订单数",
    "ord_refund_amt": "下单口径退款金额",
    "ord_refund_ord_cnt": "下单口径退款订单数",
    "ord_confirm_amt": "下单口径收货金额",
    "ord_confirm_ord_cnt": "下单口径收货订单数",
    "pay_refund_amt": "支付口径退款金额",
    "pay_refund_ord_cnt": "支付口径退款订单数",
    "pay_confirm_amt": "支付口径收货金额",
    "pay_confirm_ord_cnt": "支付口径收货订单数",
    "reward_byr_cnt": "打赏人数",
    "reward_cnt": "打赏次数",
    "reward_gift_cnt": "礼物打赏次数",
    "reward_taohuahua_cnt": "淘花花打赏次数",
    "reward_rate": "打赏率",
    "rank": "排名",
    "ranking": "排名",
    "rn": "排名",
    "row_num": "排名",
    "row_number": "排名",
    "row_count": "数据行数",
    "current_value": "当前值",
    "baseline_value": "基准值",
    "absolute_change": "绝对变化",
    "relative_change": "变化率",
    "exact_decimal": "高精度小数",
    "huge_count": "大整数值",
    "note": "备注",
    "=danger": "风险字段",
}

_export_lock = threading.RLock()


@contextmanager
def _artifact_storage_lock(storage_root: Path):
    """Serialize shared-volume quota, cleanup, and artifact writes across processes."""

    storage_root.mkdir(parents=True, exist_ok=True, mode=0o700)
    lock_path = storage_root / ".xpd-report-storage.lock"
    flags = os.O_RDWR | os.O_CREAT
    if hasattr(os, "O_CLOEXEC"):
        flags |= os.O_CLOEXEC
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    with _export_lock:
        descriptor = os.open(lock_path, flags, 0o600)
        try:
            fcntl.flock(descriptor, fcntl.LOCK_EX)
            yield
        finally:
            fcntl.flock(descriptor, fcntl.LOCK_UN)
            os.close(descriptor)


EXPORT_REPORT_FILE_SCHEMA = {
    "name": "export_report_file",
    "description": (
        "Generate a downloadable CSV, real XLSX workbook, Markdown, PDF, or JSON business "
        "report from an exact, short-lived query snapshot. Use only when the user explicitly "
        "asks to export or download a file. For a pure follow-up export, reuse the most recent "
        "result_id from this session and do not call any db_* tool again."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "format": {
                "type": "string",
                "enum": ["csv", "xlsx", "markdown", "pdf", "json"],
                "description": (
                    "Output file format. Use xlsx for an editable workbook, pdf for a "
                    "fixed-layout business report, or json for system integration."
                ),
            },
            "filename": {
                "type": "string",
                "description": "Download filename. The correct extension is added automatically.",
            },
            "title": {
                "type": "string",
                "description": "Short Chinese report title.",
            },
            "result_id": {
                "type": "string",
                "description": (
                    "The opaque result_id returned by a successful db_execute_sql call in this "
                    "session. It may come from an earlier turn and prevents exporting invented "
                    "or unexecuted data."
                ),
            },
            "analysis_type": {
                "type": "string",
                "enum": ["simple", "comparison", "diagnostic"],
                "description": (
                    "Required for XLSX. simple creates a compact query workbook; comparison "
                    "adds trend/comparison analysis; diagnostic also adds merchant-facing "
                    "anomaly, driver, and recommendation analysis. "
                    "If the user has not chosen and context is ambiguous, ask before exporting."
                ),
            },
            "analysis": {
                "type": "object",
                "description": (
                    "Evidence-backed analysis payload used to build merchant-facing XLSX summary, "
                    "trend, anomaly/recommendation, detail, and caliber sheets. Internal SQL and "
                    "database audit fields are never included in the merchant workbook."
                ),
                "properties": {
                    "metrics": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "value": {},
                                "unit": {"type": ["string", "null"]},
                                "description": {"type": ["string", "null"]},
                            },
                            "required": ["name", "value"],
                            "additionalProperties": False,
                        },
                    },
                    "comparisons": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "metric": {"type": "string"},
                                "current_value": {},
                                "baseline_value": {},
                                "absolute_change": {},
                                "relative_change": {},
                                "unit": {"type": ["string", "null"]},
                                "baseline_label": {"type": ["string", "null"]},
                                "favorability": {"type": ["string", "null"]},
                            },
                            "required": [
                                "metric",
                                "current_value",
                                "baseline_value",
                                "absolute_change",
                                "relative_change",
                            ],
                            "additionalProperties": False,
                        },
                    },
                    "trends": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "period": {"type": "string"},
                                "metric": {"type": "string"},
                                "current_value": {},
                                "baseline_value": {},
                                "absolute_change": {},
                                "relative_change": {},
                                "unit": {"type": ["string", "null"]},
                                "anomaly": {"type": ["string", "null"]},
                            },
                            "required": ["period", "metric", "current_value"],
                            "additionalProperties": False,
                        },
                    },
                    "drivers": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "statement": {"type": "string"},
                                "evidence": {"type": "string"},
                                "metric_refs": {"type": "array", "items": {"type": "string"}},
                                "contribution": {},
                                "confidence": {},
                                "dimension": {"type": ["string", "null"]},
                                "member": {"type": ["string", "null"]},
                                "metric": {"type": ["string", "null"]},
                                "current_value": {},
                                "contribution_value": {},
                                "contribution_rate": {},
                            },
                            "required": ["statement", "evidence"],
                            "additionalProperties": False,
                        },
                    },
                    "anomalies": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "statement": {"type": "string"},
                                "severity": {"type": ["string", "null"]},
                                "metric_refs": {
                                    "type": "array",
                                    "items": {"type": "string"},
                                },
                                "observed_value": {},
                                "baseline_value": {},
                            },
                            "required": ["statement"],
                            "additionalProperties": False,
                        },
                    },
                    "recommendations": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "action": {"type": "string"},
                                "rationale": {"type": "string"},
                                "priority": {},
                                "metric_refs": {"type": "array", "items": {"type": "string"}},
                            },
                            "required": ["action", "rationale"],
                            "additionalProperties": False,
                        },
                    },
                    "metric_definitions": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "formula": {"type": "string"},
                                "unit": {"type": ["string", "null"]},
                                "aggregation": {"type": ["string", "null"]},
                                "numerator": {"type": ["string", "null"]},
                                "denominator": {"type": ["string", "null"]},
                                "grain": {"type": ["string", "null"]},
                            },
                            "required": ["name", "formula"],
                            "additionalProperties": False,
                        },
                    },
                    "data_scope": {
                        "type": "object",
                        "properties": {
                            "period_start": {},
                            "period_end": {},
                            "timezone": {},
                            "grain": {},
                            "filters": {"type": "array", "items": {"type": "string"}},
                            "dimensions": {"type": "array", "items": {"type": "string"}},
                            "source_tables": {"type": "array", "items": {"type": "string"}},
                            "deduplication": {},
                        },
                        "additionalProperties": False,
                    },
                    "data_quality": {
                        "type": "object",
                        "properties": {
                            "query_count": {},
                            "returned_row_count": {},
                            "truncated": {},
                            "empty_result": {},
                            "validation_passed": {},
                            "freshness": {},
                            "period_coverage": {},
                            "null_counts": {},
                            "zero_denominators": {},
                            "small_samples": {},
                            "dimensions": {},
                            "warnings": {"type": "array", "items": {"type": "string"}},
                            "notes": {"type": "array", "items": {"type": "string"}},
                        },
                        "additionalProperties": False,
                    },
                },
                "additionalProperties": False,
            },
            "summary": {
                "type": "string",
                "description": (
                    "Concise Chinese business conclusion for XLSX/Markdown/PDF/JSON reports."
                ),
            },
            "insights": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Key Chinese business insights supported by the query result.",
            },
            "assumptions": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Query assumptions and metric definitions.",
            },
            "notes": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Caveats and usage notes.",
            },
        },
        "required": ["format", "title", "result_id"],
    },
}


def _json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, default=str)


def _error(message: Any) -> str:
    return _json({"ok": False, "error": str(message)})


def _bounded_int(name: str, default: int, minimum: int) -> int:
    try:
        return max(minimum, int(os.getenv(name, str(default))))
    except (TypeError, ValueError):
        return default


def _storage_root() -> Path:
    configured = os.getenv("XPD_FILE_STORAGE_PATH", "").strip()
    if not configured:
        raise RuntimeError("XPD_FILE_STORAGE_PATH is not configured.")
    root = Path(configured).expanduser()
    if not root.is_absolute():
        raise RuntimeError("XPD_FILE_STORAGE_PATH must be an absolute path.")
    return root.resolve()


def _session_id(session_id: Any, task_id: Any) -> str:
    for candidate in (session_id, task_id):
        value = str(candidate or "")
        if SESSION_ID_PATTERN.fullmatch(value) and "_reflection_" not in value:
            return value
    raise ValueError("Report export requires an owned xpd session.")


def _exports_dir(session_id: str) -> Path:
    root = _storage_root()
    root.mkdir(parents=True, exist_ok=True, mode=0o700)
    os.chmod(root, 0o700)
    session_candidate = root / session_id
    if session_candidate.is_symlink():
        raise ValueError("Invalid report export session path.")
    session_root = session_candidate.resolve()
    if session_root.parent != root:
        raise ValueError("Invalid report export session path.")
    session_root.mkdir(exist_ok=True, mode=0o700)
    os.chmod(session_root, 0o700)
    exports_candidate = session_root / "exports"
    if exports_candidate.is_symlink():
        raise ValueError("Invalid report export directory.")
    exports_dir = exports_candidate.resolve()
    if exports_dir.parent != session_root:
        raise ValueError("Invalid report export directory.")
    exports_dir.mkdir(exist_ok=True, mode=0o700)
    os.chmod(exports_dir, 0o700)
    return exports_dir


def _clean_text(value: Any, *, limit: int = MAX_TEXT_CHARS) -> str:
    text = str(value or "").replace("\x00", "").strip()
    return text[:limit]


def _clean_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [text for item in value[:MAX_LIST_ITEMS] if (text := _clean_text(item, limit=2000))]


def _clean_analysis_type(value: Any, *, required: bool = False) -> str:
    normalized = str(value or "").strip().lower()
    if not normalized and not required:
        return "simple"
    if normalized not in ANALYSIS_TYPES:
        raise ValueError(
            "analysis_type must be simple, comparison, or diagnostic. "
            "Ask the user to choose when the XLSX report type is unclear."
        )
    return normalized


def _clean_analysis_value(
    value: Any,
    *,
    depth: int = 0,
    list_limit: int = MAX_LIST_ITEMS,
) -> Any:
    if depth > 4:
        return None
    if value is None or isinstance(value, (bool, int, float, Decimal, date, datetime)):
        return value
    if isinstance(value, str):
        return _clean_text(value, limit=4000)
    if isinstance(value, list):
        return [
            cleaned
            for item in value[:list_limit]
            if (cleaned := _clean_analysis_value(item, depth=depth + 1)) is not None
        ]
    if isinstance(value, dict):
        cleaned_mapping: dict[str, Any] = {}
        for raw_key, raw_value in list(value.items())[:MAX_COLUMNS]:
            key = _clean_text(raw_key, limit=100)
            if not key:
                continue
            cleaned_mapping[key] = _clean_analysis_value(raw_value, depth=depth + 1)
        return cleaned_mapping
    return _clean_text(value, limit=4000)


def _analysis_record_value(record: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = record.get(key)
        if value not in (None, ""):
            return value
    return None


def _analysis_value_or(value: Any, fallback: Any) -> Any:
    return fallback if value in (None, "") else value


def _analysis_metric_value(value: Any, metric: Any, unit: Any = None) -> tuple[Any, str]:
    unit_text = _clean_text(unit, limit=40)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.endswith("%"):
            try:
                return float(stripped[:-1].replace(",", "")) / 100, unit_text or "%"
            except ValueError:
                pass
    metric_text = _clean_text(metric, limit=160)
    if not unit_text:
        if "率" in metric_text or "占比" in metric_text:
            unit_text = "%"
        elif "金额" in metric_text:
            unit_text = "元"
        elif "件数" in metric_text or "数量" in metric_text:
            unit_text = "件"
        elif "订单" in metric_text:
            unit_text = "单"
        elif "买家" in metric_text or "人数" in metric_text:
            unit_text = "人"
    return value, unit_text


def _analysis_metric_label(value: Any) -> str:
    text = _clean_text(value, limit=160)
    return _column_label(text) if text else "未命名指标"


_ANALYSIS_METRIC_PERIOD_QUALIFIERS = frozenset(
    {
        "当前周期",
        "当前期",
        "当前值",
        "本期",
        "当期",
        "基准周期",
        "基准期",
        "对比周期",
        "对比期",
        "上一周期",
        "上期",
    }
)
_ANALYSIS_METRIC_TRAILING_QUALIFIER = re.compile(r"\(([^()]*)\)\s*$")


def _analysis_metric_match_key(value: Any, unit: Any) -> tuple[str, str]:
    """Return a conservative key used only to associate KPI and comparison records."""

    text = unicodedata.normalize("NFKC", _clean_text(value, limit=160)).strip()
    unit_text = unicodedata.normalize("NFKC", _clean_text(unit, limit=40)).strip()
    while match := _ANALYSIS_METRIC_TRAILING_QUALIFIER.search(text):
        qualifier = re.sub(r"\s+", "", match.group(1))
        normalized_unit = re.sub(r"\s+", "", unit_text)
        if qualifier not in _ANALYSIS_METRIC_PERIOD_QUALIFIERS and not (
            normalized_unit and qualifier.casefold() == normalized_unit.casefold()
        ):
            break
        text = text[: match.start()].rstrip()
    return (
        re.sub(r"\s+", "", text).casefold(),
        re.sub(r"\s+", "", unit_text).casefold(),
    )


def _matching_comparison(
    metric: dict[str, Any], comparisons: list[dict[str, Any]]
) -> dict[str, Any]:
    name = _analysis_record_value(metric, "name", "metric")
    for item in comparisons:
        if _cell_text(_analysis_record_value(item, "metric", "name")) == _cell_text(name):
            return item

    match_key = _analysis_metric_match_key(name, metric.get("unit"))
    if not match_key[0]:
        return {}
    normalized_matches = [
        item
        for item in comparisons
        if _analysis_metric_match_key(
            _analysis_record_value(item, "metric", "name"), item.get("unit")
        )
        == match_key
    ]
    return normalized_matches[0] if len(normalized_matches) == 1 else {}


def _normalize_analysis(analysis: dict[str, Any]) -> dict[str, Any]:
    """Normalize canonical and legacy Agent payloads into the XLSX table contract."""

    normalized = dict(analysis)

    metrics: list[dict[str, Any]] = []
    for item in analysis.get("metrics", []) if isinstance(analysis.get("metrics"), list) else []:
        if not isinstance(item, dict):
            continue
        name = _analysis_record_value(item, "name", "metric")
        value = _analysis_record_value(item, "current_value", "value")
        if name is None or value is None:
            continue
        display_name = _analysis_metric_label(name)
        metric_value, unit = _analysis_metric_value(value, display_name, item.get("unit"))
        metrics.append(
            {
                **item,
                "name": display_name,
                "value": metric_value,
                "unit": unit,
            }
        )
    normalized["metrics"] = metrics[:MAX_LIST_ITEMS]

    comparison_meta_keys = {
        "dimension",
        "period",
        "date",
        "label",
        "session",
        "name",
        "note",
        "anomaly",
        "favorability",
        "favorable_unfavorable",
    }
    comparisons: list[dict[str, Any]] = []
    comparison_source = (
        analysis.get("comparisons", []) if isinstance(analysis.get("comparisons"), list) else []
    )
    for item in comparison_source:
        if not isinstance(item, dict):
            continue
        metric = _analysis_record_value(item, "metric", "name")
        current_value = _analysis_record_value(item, "current_value", "value")
        if metric is not None and current_value is not None:
            display_metric = _analysis_metric_label(metric)
            metric_value, unit = _analysis_metric_value(
                current_value, display_metric, item.get("unit")
            )
            comparisons.append(
                {
                    **item,
                    "metric": display_metric,
                    "current_value": metric_value,
                    "unit": unit,
                }
            )
            continue

        comparison_object = _analysis_record_value(
            item, "dimension", "session", "period", "date", "label", "name"
        )
        for key, value in item.items():
            if (
                key in comparison_meta_keys
                or value in (None, "")
                or isinstance(value, (dict, list))
            ):
                continue
            display_metric = _analysis_metric_label(key)
            metric_label = (
                f"{comparison_object}｜{display_metric}" if comparison_object else display_metric
            )
            metric_value, unit = _analysis_metric_value(value, display_metric)
            comparisons.append(
                {
                    "metric": metric_label,
                    "current_value": metric_value,
                    "baseline_label": "场次横向对比",
                    "unit": unit,
                    "favorability": "未判定",
                }
            )
    normalized["comparisons"] = comparisons[:MAX_LIST_ITEMS]

    trend_meta_keys = {
        "period",
        "date",
        "label",
        "session",
        "dimension",
        "member",
        "name",
        "note",
        "anomaly",
        "unit",
    }
    trends: list[dict[str, Any]] = []
    trend_source = analysis.get("trends", []) if isinstance(analysis.get("trends"), list) else []
    for item in trend_source:
        if not isinstance(item, dict):
            continue
        period = _analysis_record_value(item, "period", "date", "label")
        metric = _analysis_record_value(item, "metric", "name")
        current_value = _analysis_record_value(item, "current_value", "value")
        if period is not None and metric is not None and current_value is not None:
            display_metric = _analysis_metric_label(metric)
            metric_value, unit = _analysis_metric_value(
                current_value, display_metric, item.get("unit")
            )
            trends.append(
                {
                    **item,
                    "period": period,
                    "metric": display_metric,
                    "current_value": metric_value,
                    "unit": unit,
                }
            )
            continue

        session = _analysis_record_value(item, "session", "dimension", "member", "name")
        period_label = "｜".join(
            part
            for part in (_clean_text(period, limit=120), _clean_text(session, limit=120))
            if part
        )
        anomaly = _analysis_record_value(item, "anomaly", "note")
        for key, value in item.items():
            if key in trend_meta_keys or value in (None, "") or isinstance(value, (dict, list)):
                continue
            display_metric = _analysis_metric_label(key)
            metric_value, unit = _analysis_metric_value(value, display_metric)
            trends.append(
                {
                    "period": period_label or "未提供周期",
                    "metric": display_metric,
                    "current_value": metric_value,
                    "unit": unit,
                    "anomaly": anomaly,
                }
            )
    normalized["trends"] = trends[:MAX_TREND_ITEMS]

    drivers: list[dict[str, Any]] = []
    driver_source = analysis.get("drivers", []) if isinstance(analysis.get("drivers"), list) else []
    for item in driver_source:
        if not isinstance(item, dict):
            continue
        statement = _analysis_record_value(item, "statement", "driver", "name", "member")
        evidence = _analysis_record_value(item, "evidence")
        if statement is None and evidence is None:
            continue
        metric_refs = item.get("metric_refs") if isinstance(item.get("metric_refs"), list) else []
        metric = _analysis_record_value(item, "metric") or "、".join(
            _clean_text(value, limit=120) for value in metric_refs if _clean_text(value, limit=120)
        )
        impact = _analysis_record_value(item, "impact")
        current_value = _analysis_record_value(item, "current_value", "value")
        contribution_value = _analysis_record_value(item, "contribution_value")
        contribution_rate = _analysis_record_value(item, "contribution_rate", "contribution")
        confidence = _analysis_record_value(item, "confidence")
        drivers.append(
            {
                **item,
                "dimension": _analysis_record_value(item, "dimension") or impact,
                "member": _analysis_record_value(item, "member", "name", "driver")
                or statement,
                "metric": _analysis_metric_label(metric) if metric else None,
                "current_value": current_value,
                "contribution_value": contribution_value,
                "contribution_rate": contribution_rate,
                "statement": statement,
                "evidence": evidence,
                "confidence": confidence,
            }
        )
    normalized["drivers"] = drivers[:MAX_LIST_ITEMS]

    anomalies: list[dict[str, Any]] = []
    anomaly_source = (
        analysis.get("anomalies", []) if isinstance(analysis.get("anomalies"), list) else []
    )
    for item in anomaly_source:
        if not isinstance(item, dict):
            continue
        statement = _analysis_record_value(item, "statement", "anomaly", "name")
        if statement is None:
            continue
        metric_refs = item.get("metric_refs") if isinstance(item.get("metric_refs"), list) else []
        anomalies.append(
            {
                **item,
                "statement": statement,
                "metric_refs": [
                    _analysis_metric_label(value)
                    for value in metric_refs
                    if _clean_text(value, limit=120)
                ],
            }
        )
    normalized["anomalies"] = anomalies[:MAX_LIST_ITEMS]

    recommendations: list[dict[str, Any]] = []
    recommendation_source = (
        analysis.get("recommendations", [])
        if isinstance(analysis.get("recommendations"), list)
        else []
    )
    for item in recommendation_source:
        if not isinstance(item, dict):
            continue
        action = _analysis_record_value(item, "action", "statement", "recommendation")
        if action is None:
            continue
        recommendations.append(
            {
                **item,
                "action": action,
                "rationale": _analysis_record_value(item, "rationale", "evidence", "basis")
                or "依据本报告数据",
            }
        )
    normalized["recommendations"] = recommendations[:MAX_LIST_ITEMS]

    metric_definitions: list[dict[str, Any]] = []
    definition_source = (
        analysis.get("metric_definitions", [])
        if isinstance(analysis.get("metric_definitions"), list)
        else []
    )
    for item in definition_source:
        if not isinstance(item, dict):
            continue
        name = _analysis_record_value(item, "name", "metric")
        formula = _analysis_record_value(item, "formula", "definition", "description")
        if name is None and formula is None:
            continue
        metric_definitions.append(
            {
                **item,
                "name": _analysis_metric_label(name) if name else "未命名指标",
                "formula": formula or "未提供计算说明",
            }
        )
    normalized["metric_definitions"] = metric_definitions[:MAX_LIST_ITEMS]
    return normalized


def _clean_analysis(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        return {}
    allowed = {
        "metrics",
        "comparisons",
        "trends",
        "drivers",
        "anomalies",
        "recommendations",
        "metric_definitions",
        "data_scope",
        "data_quality",
    }
    cleaned = {
        key: _clean_analysis_value(
            raw_value,
            list_limit=MAX_TREND_ITEMS if key == "trends" else MAX_LIST_ITEMS,
        )
        for key, raw_value in value.items()
        if key in allowed
    }
    return _normalize_analysis(cleaned)


def _sanitize_filename(raw: Any, title: str, output_format: str) -> str:
    extension = FORMAT_EXTENSIONS[output_format]
    candidate = _clean_text(raw, limit=180) or title or "经营报告"
    candidate = unicodedata.normalize("NFKC", Path(candidate).name)
    candidate = re.sub(r"[\x00-\x1f\x7f/\\:*?\"<>|]+", "_", candidate).strip(" ._")
    if not candidate:
        candidate = "经营报告"
    for known_extension in FORMAT_EXTENSIONS.values():
        if candidate.lower().endswith(known_extension):
            candidate = candidate[: -len(known_extension)].rstrip(" ._")
            break
    candidate = candidate[:120].rstrip(" ._") or "经营报告"
    byte_budget = MAX_FILENAME_BYTES - len(extension.encode("utf-8"))
    encoded = candidate.encode("utf-8")
    if len(encoded) > byte_budget:
        candidate = encoded[:byte_budget].decode("utf-8", errors="ignore").rstrip(" ._")
    candidate = candidate or "经营报告"
    return f"{candidate}{extension}"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    return str(value)


def _xlsx_display_value(value: Any) -> Any:
    """Localize exact UI-like scalar values while preserving business text."""
    if isinstance(value, bool):
        return "是" if value else "否"
    if not isinstance(value, str):
        return value
    localized = {
        "PASS": "通过",
        "PASSED": "通过",
        "WARNING": "警告",
        "WARN": "警告",
        "FAIL": "失败",
        "FAILED": "失败",
        "TRUE": "是",
        "FALSE": "否",
        "LOW": "低",
        "MEDIUM": "中",
        "HIGH": "高",
        "F": "有利",
        "FAVORABLE": "有利",
        "U": "不利",
        "UNFAVORABLE": "不利",
    }
    return localized.get(value.strip().upper(), value)


def _xlsx_detail_display_value(column: str, value: Any) -> Any:
    normalized_column = str(column).strip().strip("`\"'").lower()
    if normalized_column in {"period_type", "comparison_period"} and isinstance(value, str):
        localized_period = {
            "CURRENT": "当前周期",
            "CURRENT_PERIOD": "当前周期",
            "BASELINE": "基准周期",
            "BASELINE_PERIOD": "基准周期",
            "PREVIOUS": "基准周期",
            "PREVIOUS_PERIOD": "基准周期",
        }
        if localized := localized_period.get(value.strip().upper()):
            return localized
    return _xlsx_display_value(value)


def _column_label(column: str) -> str:
    raw = str(column).strip()
    if re.search(r"[\u3400-\u9fff]", raw):
        return raw
    normalized = raw.strip("`\"'").lower()

    def known_label(identifier: str) -> str | None:
        if identifier in CHINESE_COLUMN_LABELS:
            return CHINESE_COLUMN_LABELS[identifier]

        period_prefixes = (
            ("current_period_", "当前周期"),
            ("current_", "当前周期"),
            ("baseline_period_", "基准周期"),
            ("baseline_", "基准周期"),
            ("previous_period_", "基准周期"),
            ("previous_", "基准周期"),
            ("prev_period_", "基准周期"),
            ("prev_", "基准周期"),
        )
        for prefix, label in period_prefixes:
            if identifier.startswith(prefix):
                remainder_label = known_label(identifier[len(prefix) :])
                if remainder_label:
                    return f"{label}{remainder_label}"

        period_suffixes = (
            ("_current_period", "当前周期"),
            ("_current", "当前周期"),
            ("_baseline_period", "基准周期"),
            ("_baseline", "基准周期"),
            ("_previous_period", "基准周期"),
            ("_previous", "基准周期"),
            ("_prev_period", "基准周期"),
            ("_prev", "基准周期"),
        )
        for suffix, label in period_suffixes:
            if identifier.endswith(suffix):
                remainder_label = known_label(identifier[: -len(suffix)])
                if remainder_label:
                    return f"{label}{remainder_label}"

        change_prefixes = (
            ("absolute_change_", "绝对变化"),
            ("abs_change_", "绝对变化"),
            ("change_", "绝对变化"),
            ("delta_", "绝对变化"),
            ("diff_", "绝对变化"),
        )
        for prefix, label in change_prefixes:
            if identifier.startswith(prefix):
                remainder_label = known_label(identifier[len(prefix) :])
                if remainder_label:
                    return f"{remainder_label}{label}"

        change_suffixes = (
            ("_absolute_change", "绝对变化"),
            ("_abs_change", "绝对变化"),
            ("_change", "绝对变化"),
            ("_delta", "绝对变化"),
            ("_diff", "绝对变化"),
        )
        for suffix, label in change_suffixes:
            if identifier.endswith(suffix):
                remainder_label = known_label(identifier[: -len(suffix)])
                if remainder_label:
                    return f"{remainder_label}{label}"

        aggregate_prefixes = (
            ("total_", "总"),
            ("sum_", "合计"),
            ("avg_", "平均"),
            ("average_", "平均"),
            ("max_", "最高"),
            ("min_", "最低"),
        )
        for prefix, label in aggregate_prefixes:
            if identifier.startswith(prefix):
                remainder_label = known_label(identifier[len(prefix) :])
                if remainder_label:
                    return f"{label}{remainder_label}"

        aggregate_suffixes = (
            ("_total", "合计"),
            ("_sum", "合计"),
            ("_avg", "平均值"),
            ("_average", "平均值"),
            ("_max", "最高值"),
            ("_min", "最低值"),
        )
        for suffix, label in aggregate_suffixes:
            if identifier.endswith(suffix):
                remainder_label = known_label(identifier[: -len(suffix)])
                if remainder_label:
                    return f"{remainder_label}{label}"
        return None

    if label := known_label(normalized):
        return label

    # Keep unknown technical identifiers in the cell comment only.  A raw SQL
    # alias must not leak back into the user-facing Chinese header.
    return "其他数据字段"


def _display_columns(columns: list[str]) -> list[str]:
    labels: list[str] = []
    counts: dict[str, int] = {}
    for column in columns:
        label = _column_label(column)
        counts[label] = counts.get(label, 0) + 1
        count = counts[label]
        labels.append(label if count == 1 else f"{label}（{count}）")
    return labels


def _merchant_detail_columns(columns: list[str]) -> list[str]:
    """Keep the exact query fields while placing optional identifiers last."""

    business_columns = [column for column in columns if not _is_identifier_column(column)]
    identifier_columns = [column for column in columns if _is_identifier_column(column)]
    return business_columns + identifier_columns


def _csv_cell(value: Any) -> str:
    text = _cell_text(value)
    if isinstance(value, str) and text.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _csv_bytes(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    # SQL aliases can become column names, so headers need the same spreadsheet
    # formula-injection protection as data cells.
    writer.writerow([_csv_cell(column) for column in _display_columns(columns)])
    for row in rows:
        writer.writerow([_csv_cell(row.get(column)) for column in columns])
    return output.getvalue().encode("utf-8-sig")


def _markdown_cell(value: Any) -> str:
    return (
        _cell_text(value)
        .replace("\\", "\\\\")
        .replace("|", "\\|")
        .replace("\r\n", "<br>")
        .replace("\n", "<br>")
        .replace("\r", "<br>")
    )


def _markdown_list(lines: list[str], heading: str, items: list[str]) -> None:
    if not items:
        return
    lines.extend(("", f"## {heading}", ""))
    lines.extend(f"- {item}" for item in items)


def _markdown_bytes(
    *,
    title: str,
    summary: str,
    insights: list[str],
    assumptions: list[str],
    notes: list[str],
    sql: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    truncated: bool,
) -> bytes:
    lines = [f"# {title}"]
    if summary:
        lines.extend(("", "## 经营结论", "", summary))
    _markdown_list(lines, "关键洞察", insights)
    lines.extend(("", "## 数据明细", ""))
    if columns:
        lines.append("| " + " | ".join(_markdown_cell(column) for column in columns) + " |")
        lines.append("| " + " | ".join("---" for _ in columns) + " |")
        lines.extend(
            "| " + " | ".join(_markdown_cell(row.get(column)) for column in columns) + " |"
            for row in rows
        )
    else:
        lines.append("（查询结果为空）")
    if truncated:
        lines.extend(("", "> 数据达到导出行数上限，本文件仅包含前若干行。"))
    _markdown_list(lines, "查询假设", assumptions)
    lines.extend(("", "## 已执行 SQL", ""))
    lines.extend(f"    {line}" for line in sql.strip().splitlines())
    _markdown_list(lines, "注意事项", notes)
    lines.append("")
    return "\n".join(lines).encode("utf-8")


def _report_timezone() -> ZoneInfo:
    name = os.getenv("HERMES_TIMEZONE", "Asia/Shanghai").strip() or "Asia/Shanghai"
    try:
        return ZoneInfo(name)
    except ZoneInfoNotFoundError as exc:
        raise ValueError(f"Unknown HERMES_TIMEZONE: {name}") from exc


def _xlsx_text(value: Any, *, context: str) -> str:
    text = _cell_text(value)
    text = EXCEL_ILLEGAL_CONTROL_PATTERN.sub(
        lambda match: f"\\x{ord(match.group(0)):02x}",
        text,
    )
    if text.lstrip().startswith(("=", "+", "-", "@")):
        text = f"'{text}"
    if len(text) > EXCEL_MAX_CELL_CHARS:
        raise ValueError(
            f"{context} contains {len(text)} characters; Excel cells support at most "
            f"{EXCEL_MAX_CELL_CHARS}. The value was not truncated."
        )
    return text


def _xlsx_safe_value(
    value: Any,
    *,
    context: str = "XLSX cell",
    force_text: bool = False,
) -> Any:
    """Return an Excel-safe value without silently losing source precision or text."""
    if value is None:
        return None
    if force_text:
        return _xlsx_text(value, context=context)
    if isinstance(value, bool):
        return value
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(_report_timezone()).replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, int):
        if abs(value) > EXCEL_MAX_EXACT_INTEGER:
            return _xlsx_text(value, context=context)
        return value
    if isinstance(value, Decimal):
        if value.is_finite():
            try:
                numeric = float(value)
            except (TypeError, ValueError, OverflowError):
                numeric = math.nan
            significant_digits = len(value.normalize().as_tuple().digits)
            if (
                significant_digits <= 15
                and math.isfinite(numeric)
                and Decimal(str(numeric)) == value
            ):
                return numeric
        return _xlsx_text(value, context=context)
    if isinstance(value, float):
        return value if math.isfinite(value) else _xlsx_text(value, context=context)
    return _xlsx_text(value, context=context)


def _display_width(value: Any) -> int:
    return sum(
        2 if re.match(r"[\u3400-\u9fff]", character) else 1 for character in _cell_text(value)
    )


def _wrapped_line_count(value: Any, width: float) -> int:
    available_width = max(1, int(width) - 2)
    physical_lines = re.split(r"\r\n?|\n", _cell_text(value))
    return sum(max(1, math.ceil(_display_width(line) / available_width)) for line in physical_lines)


def _is_identifier_column(column: str) -> bool:
    raw = str(column).strip("`\"'").strip()
    normalized = raw.lower()
    return (
        normalized == "id"
        or normalized.endswith("_id")
        or _column_label(raw).upper().endswith("ID")
    )


def _column_widths(columns: list[str], rows: list[dict[str, Any]]) -> list[float]:
    widths = []
    labels = _display_columns(columns)
    for column, label in zip(columns, labels, strict=True):
        normalized = str(column).strip("`\"'").lower()
        longest = _display_width(label)
        for row in rows[:200]:
            longest = max(longest, _display_width(row.get(column)))
        if any(token in normalized for token in ("time", "date")):
            minimum, maximum = 18, 22
        elif any(token in normalized for token in ("title", "name", "caliber", "desc")):
            minimum, maximum = 20, 32
        elif normalized.endswith("_id"):
            minimum, maximum = 18, 24
        elif normalized.endswith(
            (
                "_amt",
                "_amount",
                "_price",
                "_cost",
                "_rate",
                "_ratio",
                "_pct",
                "_percentage",
                "_cnt",
                "_count",
                "_qty",
                "_pv",
                "_uv",
            )
        ):
            minimum, maximum = 12, 16
        else:
            minimum, maximum = 12, 22
        widths.append(float(min(maximum, max(minimum, longest + 2))))
    return widths or [16.0]


def _column_number_kinds(columns: list[str]) -> list[str | None]:
    kinds: list[str | None] = []
    for column in columns:
        normalized = str(column).strip("`\"'").lower()
        label = _column_label(str(column))
        if _is_identifier_column(str(column)):
            kinds.append(None)
        elif normalized.endswith(("_pct", "_percent", "_percentage_point")):
            kinds.append("percent_point")
        elif (
            normalized.endswith(("_rate", "_ratio", "_pct", "_percentage"))
            or "率" in label
            or "占比" in label
        ):
            kinds.append("percent")
        elif normalized.endswith(("_amt", "_amount", "_price", "_cost")) or any(
            token in label for token in ("金额", "单价", "均价", "成本")
        ):
            kinds.append("currency")
        elif (
            normalized.endswith(("_cnt", "_count", "_qty", "_pv", "_uv"))
            or normalized in {"rank", "ranking", "row_count", "live_duration_min"}
            or any(token in label for token in ("人数", "次数", "件数", "订单数", "排名"))
        ):
            kinds.append("integer")
        elif normalized in {
            "total_hours",
            "duration_hours",
            "avg_watch_sec_per_user",
            "avg_watch_sec_per_view",
        }:
            kinds.append("decimal")
        else:
            kinds.append(None)
    return kinds


def _xlsx_bytes(
    *,
    title: str,
    summary: str,
    insights: list[str],
    assumptions: list[str],
    notes: list[str],
    sql: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    truncated: bool,
    analysis_type: str = "simple",
    analysis: dict[str, Any] | None = None,
) -> bytes:
    """Build a controlled, fixed-template report workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("XLSX export requires the openpyxl package.") from exc

    navy = "1F4E78"
    teal = "D9E2F3"
    core_blue = "1F4E78"
    chart_secondary = "A6A6A6"
    accent = navy
    text_color = "222222"
    muted = "666666"
    light_teal = "F2F2F2"
    banded = "FAFAFA"
    border_color = "D9D9D9"
    warning_fill = "FFF2CC"
    warning_text = "9C6500"
    pass_fill = "E2F0D9"
    pass_text = "548235"
    fail_fill = "FCE4D6"
    fail_text = "C00000"
    blue_fill = "FFFFFF"
    white = "FFFFFF"

    thin = Side(style="thin", color=border_color)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bottom_border = Border(bottom=thin)
    subtotal_border = Border(top=Side(style="medium", color=navy))
    font_name = "Microsoft YaHei"
    body_font = Font(name="KaiTi", size=11, color=text_color)
    label_font = Font(name=font_name, size=10, bold=True, color=accent)
    header_font = Font(name=font_name, size=10, bold=True, color=navy)
    title_font = Font(name=font_name, size=18, bold=True, color=white)
    section_font = Font(name=font_name, size=11, bold=True, color=navy)

    analysis_type = _clean_analysis_type(analysis_type)
    analysis = _normalize_analysis(dict(analysis or {}))
    metrics = [item for item in analysis.get("metrics", []) if isinstance(item, dict)][:12]
    comparisons = [item for item in analysis.get("comparisons", []) if isinstance(item, dict)][
        :MAX_LIST_ITEMS
    ]
    trends = [item for item in analysis.get("trends", []) if isinstance(item, dict)][
        :MAX_TREND_ITEMS
    ]
    drivers = [item for item in analysis.get("drivers", []) if isinstance(item, dict)][
        :MAX_LIST_ITEMS
    ]
    anomalies = [item for item in analysis.get("anomalies", []) if isinstance(item, dict)][
        :MAX_LIST_ITEMS
    ]
    recommendations = [
        item for item in analysis.get("recommendations", []) if isinstance(item, dict)
    ][:MAX_LIST_ITEMS]
    metric_definitions = [
        item for item in analysis.get("metric_definitions", []) if isinstance(item, dict)
    ][:MAX_LIST_ITEMS]
    data_scope = analysis.get("data_scope") if isinstance(analysis.get("data_scope"), dict) else {}
    data_quality = (
        analysis.get("data_quality") if isinstance(analysis.get("data_quality"), dict) else {}
    )

    def analysis_label() -> str:
        return {
            "simple": "简单查询",
            "comparison": "对比分析",
            "diagnostic": "诊断分析",
        }[analysis_type]

    def record_value(record: dict[str, Any], *keys: str) -> Any:
        for key in keys:
            value = record.get(key)
            if value not in (None, ""):
                return value
        return None

    def display_analysis_value(value: Any) -> Any:
        if isinstance(value, list):
            return "、".join(_cell_text(_xlsx_display_value(item)) for item in value)
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False, default=str)
        return _xlsx_display_value(value)

    def analysis_cell_value(value: Any, fallback: str = "未提供") -> Any:
        displayed = display_analysis_value(value)
        return fallback if displayed in (None, "") else displayed

    def favorability_label(value: Any) -> str:
        normalized = _cell_text(value).strip().upper()
        if normalized in {"F", "FAVORABLE", "有利"}:
            return "有利"
        if normalized in {"U", "UNFAVORABLE", "不利"}:
            return "不利"
        return "暂不判断"

    def flatten_mapping(value: dict[str, Any], prefix: str = "") -> list[tuple[str, Any]]:
        flattened: list[tuple[str, Any]] = []
        for key, item in value.items():
            label = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(item, dict):
                flattened.extend(flatten_mapping(item, label))
            else:
                flattened.append((label, display_analysis_value(item)))
        return flattened[:MAX_COLUMNS]

    field_labels = {
        "requested_period": "请求统计周期",
        "observed_period": "实际覆盖周期",
        "session_count": "直播场次数",
        "sessions": "直播场次数",
        "covered_days": "已覆盖天数",
        "expected_days": "应覆盖天数",
        "coverage_ratio": "周期覆盖率",
        "product_table_coverage": "商品明细覆盖范围",
        "granularity": "数据粒度",
        "actual_period": "实际数据周期",
        "baseline": "对比基准",
        "table": "数据表",
        "source_table": "数据表",
        "total_anomaly_items": "异常商品总数",
        "period_start": "统计开始日期",
        "period_end": "统计结束日期",
        "timezone": "时区",
        "grain": "数据粒度",
        "filters": "筛选条件",
        "dimensions": "分析维度",
        "source_tables": "数据表",
        "deduplication": "去重方式",
        "query_count": "查询次数",
        "scope": "检查范围",
        "returned_row_count": "返回行数",
        "column_count": "字段数",
        "max_rows": "最大返回行数",
        "truncated": "结果是否截断",
        "empty_result": "结果是否为空",
        "validation_passed": "校验是否通过",
        "freshness": "数据时效",
        "coverage": "周期覆盖情况",
        "refund_settlement": "退款沉淀情况",
        "caliber_gap": "口径差异",
        "small_sample": "小样本提示",
        "missing_dimensions": "缺失分析维度",
        "cross_period_note": "跨期退款说明",
        "null_handling": "空值处理",
        "period_coverage": "周期覆盖情况",
        "truncation": "明细截断说明",
        "cross_caliber": "跨口径说明",
        "dedup": "去重说明",
        "notes": "质量备注",
        "freshness.latest": "数据最新日期",
        "freshness.field": "数据时间字段",
        "freshness.lag_days": "数据延迟天数",
        "period_coverage.requested_start": "请求周期开始日期",
        "period_coverage.requested_end": "请求周期结束日期",
        "period_coverage.observed_start": "实际数据开始日期",
        "period_coverage.observed_end": "实际数据结束日期",
        "period_coverage.covered_days": "已覆盖天数",
        "period_coverage.expected_days": "应覆盖天数",
        "period_coverage.coverage_ratio": "周期覆盖率",
        "period_coverage.complete": "周期是否完整",
        "small_samples.threshold": "小样本阈值",
        "dimensions.required": "必要维度",
        "dimensions.present": "已有维度",
        "dimensions.missing": "缺失维度",
        "warnings": "质量警告",
        "affected_rows": "受影响行数",
        "minimum": "最小值",
    }

    def analysis_field_label(key: str) -> str:
        if key in field_labels:
            return field_labels[key]
        if key.startswith("null_counts."):
            return f"空值数量：{_column_label(key.removeprefix('null_counts.'))}"
        if key.startswith("zero_denominators."):
            return f"零分母数量：{_column_label(key.removeprefix('zero_denominators.'))}"
        if key.startswith("small_samples.columns."):
            return f"小样本字段：{_column_label(key.removeprefix('small_samples.columns.'))}"
        if re.search(r"[\u3400-\u9fff]", key):
            return key
        return "其他补充说明"

    technical_expression_pattern = re.compile(
        r"\b(?:SELECT|FROM|WHERE|JOIN|SUM|COUNT|AVG|MIN|MAX|NULLIF|CASE|WHEN|THEN)\b"
        r"|[`]|\b[A-Za-z][A-Za-z0-9]*_[A-Za-z0-9_]+\b",
        re.IGNORECASE,
    )

    def merchant_text(value: Any) -> str | None:
        displayed = display_analysis_value(value)
        if displayed in (None, ""):
            return None
        text = _cell_text(displayed).strip()
        if not text or technical_expression_pattern.search(text):
            return None
        return text

    def metric_label_with_unit(record: dict[str, Any]) -> str:
        name = analysis_cell_value(record_value(record, "metric", "name"), "未命名指标")
        unit = _cell_text(record_value(record, "unit")).strip()
        if unit and f"（{unit}）" not in _cell_text(name):
            return f"{name}（{unit}）"
        return _cell_text(name)

    def metric_definition_text(record: dict[str, Any]) -> str:
        for key in ("description", "definition", "caliber"):
            if text := merchant_text(record.get(key)):
                return text
        numerator = record_value(record, "numerator")
        denominator = record_value(record, "denominator")
        if numerator is not None and denominator is not None:
            numerator_label = _column_label(_cell_text(numerator))
            denominator_label = _column_label(_cell_text(denominator))
            if not {
                "其他数据字段",
                "补充数据",
            } & {numerator_label, denominator_label}:
                return f"{numerator_label} ÷ {denominator_label}"
        if text := merchant_text(record_value(record, "formula")):
            return text
        metric_name = _cell_text(record_value(record, "name", "metric"))
        if "件数" in metric_name and "率" in metric_name:
            return "退款商品件数 ÷ 成交商品件数"
        if "金额" in metric_name and "率" in metric_name:
            return "退款金额 ÷ 成交金额"
        return "按本报告已确认的业务口径统计"

    def joined_unique(parts: list[Any]) -> str | None:
        values: list[str] = []
        for part in parts:
            text = _cell_text(part).strip() if part not in (None, "") else ""
            if text and text not in values:
                values.append(text)
        return "；".join(values) if values else None

    def meaningful_value(value: Any) -> Any:
        if _cell_text(value).strip() in {
            "",
            "见证据",
            "未量化",
            "待验证",
            "未提供",
            "未提供结论",
            "未提供证据",
        }:
            return None
        return value

    def format_metric_display(value: Any, unit: Any, metric_name: Any = "") -> str:
        if value is None:
            return "—"
        unit_text = _cell_text(unit)
        is_rate = unit_text in {"%", "百分比", "百分点"} or "率" in _cell_text(metric_name)
        if (
            isinstance(value, (int, float, Decimal))
            and unit_text == "百分点"
            and abs(float(value)) <= 1
        ):
            return f"{float(value) * 100:.2f}个百分点"
        if isinstance(value, (int, float, Decimal)) and is_rate and abs(float(value)) <= 1:
            return f"{float(value):.2%}"
        if isinstance(value, (int, float, Decimal)):
            return f"{float(value):,.2f}".rstrip("0").rstrip(".") + unit_text
        return f"{_cell_text(value)}{unit_text}"

    def apply_metric_number_format(cell: Any, metric_name: Any, unit: Any) -> None:
        metric_text = _cell_text(metric_name)
        unit_text = _cell_text(unit)
        if "率" in metric_text or unit_text in {"%", "百分比", "百分点"}:
            cell.number_format = "0.00%;[Red](0.00%);-"
        elif unit_text in {"元", "人民币"} or "金额" in metric_text:
            cell.number_format = "¥#,##0.00;[Red](¥#,##0.00);-"
        elif unit_text in {"件", "单", "人", "次", "场"}:
            cell.number_format = "#,##0;[Red](#,##0);-"

    def configure_sheet(worksheet: Any, *, tab_color: str = navy) -> None:
        worksheet.sheet_properties.tabColor = tab_color
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 90
        worksheet.sheet_format.defaultRowHeight = 22

    def add_sheet_title(worksheet: Any, heading: str, end_column: int) -> None:
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
        cell = worksheet.cell(1, 1, _xlsx_safe_value(heading, context=f"{heading} title"))
        cell.font = title_font
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        worksheet.row_dimensions[1].height = 40

    def write_table(
        worksheet: Any,
        *,
        start_row: int,
        headers: list[str],
        records: list[list[Any]],
        widths: list[float] | None = None,
        percent_columns: set[int] | None = None,
    ) -> int:
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(start_row, column_index, header)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor=teal)
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        worksheet.row_dimensions[start_row].height = 30
        for row_offset, record in enumerate(records, start=1):
            row_index = start_row + row_offset
            for column_index, value in enumerate(record, start=1):
                safe_value = _xlsx_safe_value(
                    display_analysis_value(value),
                    context=f"{worksheet.title} row {row_index}, column {column_index}",
                )
                cell = worksheet.cell(row_index, column_index, safe_value)
                cell.font = body_font
                cell.fill = PatternFill("solid", fgColor=banded if row_index % 2 else white)
                cell.border = cell_border
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )
                if column_index in (percent_columns or set()) and isinstance(
                    safe_value, (int, float)
                ):
                    cell.number_format = "0.0%;[Red](0.0%);-"
                elif isinstance(safe_value, (int, float)) and not isinstance(safe_value, bool):
                    cell.number_format = "#,##0.00;[Red](#,##0.00);-"
            worksheet.row_dimensions[row_index].height = 26
        for index, width in enumerate(widths or [20.0] * len(headers), start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width
        return start_row + max(1, len(records))

    workbook = Workbook()
    workbook.iso_dates = True
    workbook.properties.creator = "直播经营报告助手"
    workbook.properties.title = title
    workbook.properties.subject = "直播经营数据分析报告"
    workbook.properties.description = "由直播经营报告助手受控导出器生成"

    summary_sheet = workbook.active
    summary_sheet.title = "经营摘要"
    configure_sheet(summary_sheet, tab_color=navy)
    summary_sheet.freeze_panes = "A4"
    summary_sheet.sheet_format.defaultRowHeight = 22
    summary_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    summary_sheet.page_setup.orientation = "landscape"
    summary_sheet.page_setup.paperSize = summary_sheet.PAPERSIZE_A4
    summary_sheet.page_setup.fitToWidth = 1
    summary_sheet.page_setup.fitToHeight = 0
    summary_sheet.page_margins = PageMargins(
        left=0.35,
        right=0.35,
        top=0.55,
        bottom=0.55,
        header=0.2,
        footer=0.2,
    )
    for index, width in enumerate((9.0,) * 16, start=1):
        summary_sheet.column_dimensions[get_column_letter(index)].width = width

    summary_sheet.merge_cells("A1:P1")
    title_cell = summary_sheet["A1"]
    title_cell.value = _xlsx_safe_value(title, context="Report title")
    title_cell.font = title_font
    title_cell.fill = PatternFill("solid", fgColor=navy)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    summary_sheet.row_dimensions[1].height = 40

    generated_at = datetime.now(_report_timezone()).strftime("%Y-%m-%d %H:%M:%S")
    quality_warnings = (
        data_quality.get("warnings") if isinstance(data_quality.get("warnings"), list) else []
    )
    period_start = display_analysis_value(data_scope.get("period_start"))
    period_end = display_analysis_value(data_scope.get("period_end"))
    if period_start or period_end:
        report_period = (
            period_start
            if period_start == period_end
            else f"{period_start or '未提供'} 至 {period_end or '未提供'}"
        )
    else:
        report_period = (
            display_analysis_value(data_scope.get("requested_period"))
            or display_analysis_value(data_scope.get("observed_period"))
            or "未提供"
        )
    freshness = data_quality.get("freshness")
    data_as_of = (
        (display_analysis_value(freshness.get("latest")) if isinstance(freshness, dict) else None)
        or (display_analysis_value(freshness) if freshness is not None else None)
        or "未提供"
    )
    metadata = (
        (1, 2, 8, "报告周期", report_period, None),
        (9, 10, 16, "数据截至", data_as_of, None),
    )
    for label_column, value_column, end_column, label, value, status in metadata:
        summary_sheet.merge_cells(
            start_row=2,
            start_column=value_column,
            end_row=2,
            end_column=end_column,
        )
        label_cell = summary_sheet.cell(row=2, column=label_column, value=label)
        label_cell.font = label_font
        label_cell.fill = PatternFill("solid", fgColor=light_teal)
        label_cell.border = bottom_border
        label_cell.alignment = Alignment(vertical="center", indent=1)
        value_cell = summary_sheet.cell(
            row=2,
            column=value_column,
            value=_xlsx_safe_value(value, context=f"Summary metadata {label}"),
        )
        value_cell.font = Font(
            name=font_name,
            size=10,
            color={"PASS": pass_text, "WARNING": warning_text, "FAIL": fail_text}.get(
                status, text_color
            )
            if status
            else text_color,
            bold=bool(status),
        )
        value_cell.fill = PatternFill(
            "solid",
            fgColor={"PASS": pass_fill, "WARNING": warning_fill, "FAIL": fail_fill}.get(
                status, white
            )
            if status
            else white,
        )
        value_cell.border = bottom_border
        value_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            indent=1,
        )
    summary_sheet.row_dimensions[2].height = 30
    summary_sheet.merge_cells("A3:P3")
    status_detail_cell = summary_sheet["A3"]
    status_detail_cell.value = _xlsx_safe_value(
        f"数据行数 {len(rows):,}｜生成时间 {generated_at}",
        context="Dashboard status detail",
    )
    status_detail_cell.font = Font(name="KaiTi", size=10, color=muted)
    status_detail_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    summary_sheet.row_dimensions[3].height = 22

    current_row = 5

    if metrics:
        summary_sheet.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=16
        )
        metric_heading = summary_sheet.cell(current_row, 1, "核心指标")
        metric_heading.font = section_font
        metric_heading.fill = PatternFill("solid", fgColor=light_teal)
        metric_heading.alignment = Alignment(vertical="center", indent=1)
        summary_sheet.row_dimensions[current_row].height = 28
        current_row += 1
        for metric_index, metric in enumerate(metrics):
            card_column = 1 + (metric_index % 4) * 4
            card_row = current_row + (metric_index // 4) * 4
            end_column = card_column + 3
            summary_sheet.merge_cells(
                start_row=card_row,
                start_column=card_column,
                end_row=card_row,
                end_column=end_column,
            )
            name = record_value(metric, "name", "metric") or f"指标{metric_index + 1}"
            name_cell = summary_sheet.cell(
                card_row, card_column, _xlsx_safe_value(name, context="Metric name")
            )
            name_cell.font = label_font
            name_cell.fill = PatternFill("solid", fgColor=blue_fill)
            name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            summary_sheet.merge_cells(
                start_row=card_row + 1,
                start_column=card_column,
                end_row=card_row + 1,
                end_column=end_column,
            )
            current_value = record_value(metric, "current_value", "value")
            unit = record_value(metric, "unit") or ""
            value_text = format_metric_display(current_value, unit, name)
            value_cell = summary_sheet.cell(
                card_row + 1, card_column, _xlsx_safe_value(value_text, context="Metric value")
            )
            value_cell.font = Font(name="Aptos", size=20, bold=True, color=core_blue)
            value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            summary_sheet.merge_cells(
                start_row=card_row + 2,
                start_column=card_column,
                end_row=card_row + 2,
                end_column=end_column,
            )
            comparison_item = _matching_comparison(metric, comparisons)
            baseline = record_value(metric, "baseline_value")
            if baseline is None:
                baseline = record_value(comparison_item, "baseline_value")
            absolute_change = record_value(metric, "absolute_change")
            if absolute_change is None:
                absolute_change = record_value(comparison_item, "absolute_change")
            relative_change = record_value(metric, "relative_change")
            if relative_change is None:
                relative_change = record_value(comparison_item, "relative_change")
            context_parts = []
            if baseline is not None:
                context_parts.append(f"基准 {format_metric_display(baseline, unit, name)}")
            if absolute_change is not None:
                change_unit = "百分点" if "率" in _cell_text(name) else unit
                context_parts.append(
                    f"变化 {format_metric_display(absolute_change, change_unit, name)}"
                )
            if relative_change is not None:
                context_parts.append(
                    f"变化率 {format_metric_display(relative_change, '%', '变化率')}"
                )
            context_cell = summary_sheet.cell(
                card_row + 2,
                card_column,
                _xlsx_safe_value(
                    "｜".join(context_parts) or "暂无对比基准", context="Metric comparison"
                ),
            )
            context_cell.font = Font(name="KaiTi", size=10, color=muted)
            context_cell.alignment = Alignment(
                horizontal="left", vertical="center", indent=1, wrap_text=True
            )
            for row_number in range(card_row, card_row + 3):
                for column_number in range(card_column, end_column + 1):
                    summary_sheet.cell(row_number, column_number).border = cell_border
        current_row += math.ceil(len(metrics) / 4) * 4

    def add_summary_section(heading: str, values: list[str], *, bullets: bool) -> None:
        nonlocal current_row
        if not values:
            return
        summary_sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=16,
        )
        heading_cell = summary_sheet.cell(row=current_row, column=1, value=heading)
        heading_cell.font = section_font
        heading_cell.fill = PatternFill("solid", fgColor=light_teal)
        heading_cell.border = bottom_border
        heading_cell.alignment = Alignment(vertical="center", indent=1)
        summary_sheet.row_dimensions[current_row].height = 28
        current_row += 1

        for value in values:
            display_value = f"• {value}" if bullets else value
            summary_sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=16,
            )
            content_cell = summary_sheet.cell(
                row=current_row,
                column=1,
                value=_xlsx_safe_value(
                    display_value,
                    context=f"Summary content row {current_row}",
                ),
            )
            content_cell.font = body_font
            content_cell.border = bottom_border
            content_cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
                indent=1,
            )
            estimated_lines = _wrapped_line_count(display_value, 94)
            summary_sheet.row_dimensions[current_row].height = float(
                min(300, max(30, 18 * estimated_lines + 10))
            )
            current_row += 1
        current_row += 1

    anomaly_summaries = [
        _cell_text(record_value(item, "statement", "anomaly", "name"))
        for item in anomalies[:3]
        if record_value(item, "statement", "anomaly", "name") is not None
    ]
    driver_summaries = [
        _cell_text(record_value(item, "statement", "evidence", "member", "name"))
        for item in drivers[:3]
        if record_value(item, "statement", "evidence", "member", "name") is not None
    ]
    recommendation_summaries = [
        _cell_text(record_value(item, "action", "statement"))
        for item in recommendations[:3]
        if record_value(item, "action", "statement") is not None
    ]
    add_summary_section("核心结论", [summary] if summary else [], bullets=False)
    add_summary_section(
        "主要异常与影响因素",
        (anomaly_summaries + driver_summaries)[:3] or insights[:3],
        bullets=True,
    )
    add_summary_section("建议动作", recommendation_summaries, bullets=True)
    add_summary_section(
        "数据质量警告", [_cell_text(item) for item in quality_warnings] + notes, bullets=True
    )
    add_summary_section("数据口径", assumptions, bullets=True)
    summary_sheet.print_area = f"A1:P{max(4, current_row - 1)}"
    summary_sheet.oddFooter.center.text = "直播经营数据分析报告"
    summary_sheet.oddFooter.right.text = "第 &[Page] 页"
    summary_sheet.oddFooter.center.size = 8
    summary_sheet.oddFooter.center.color = muted
    summary_sheet.oddFooter.right.size = 8
    summary_sheet.oddFooter.right.color = muted

    if analysis_type in {"comparison", "diagnostic"} and (comparisons or trends):
        trend_sheet = workbook.create_sheet("趋势与对比")
        configure_sheet(trend_sheet)
        add_sheet_title(trend_sheet, "趋势与对比", 7)
        comparison_rows = [
            [
                metric_label_with_unit(item),
                analysis_cell_value(record_value(item, "current_value", "value")),
                analysis_cell_value(record_value(item, "baseline_value"), "暂无可比数据"),
                analysis_cell_value(record_value(item, "absolute_change"), "不适用"),
                analysis_cell_value(record_value(item, "relative_change"), "不适用"),
                analysis_cell_value(record_value(item, "baseline_label"), "暂无可比周期"),
                favorability_label(record_value(item, "favorability", "favorable_unfavorable")),
            ]
            for item in comparisons
        ]
        comparison_end = write_table(
            trend_sheet,
            start_row=3,
            headers=[
                "指标",
                "当前值",
                "基准值",
                "绝对变化",
                "变化率",
                "基准周期",
                "结果判断",
            ],
            records=comparison_rows,
            widths=[28, 16, 16, 16, 14, 24, 14],
            percent_columns={5},
        )
        for row_offset, item in enumerate(comparisons, start=1):
            metric_name = _cell_text(record_value(item, "metric", "name"))
            unit = _cell_text(record_value(item, "unit"))
            for column_index in (2, 3, 4):
                value_cell = trend_sheet.cell(3 + row_offset, column_index)
                if isinstance(value_cell.value, (int, float)):
                    apply_metric_number_format(value_cell, metric_name, unit)
            comparison_row = 3 + row_offset
            if any(
                keyword in metric_name.upper()
                for keyword in ("毛利", "CM I", "CM II", "EBITDA", "经营利润", "净利润")
            ):
                for column_index in range(1, 8):
                    cell = trend_sheet.cell(comparison_row, column_index)
                    cell.fill = PatternFill("solid", fgColor=blue_fill)
                    cell.border = subtotal_border
                    cell.font = Font(name=font_name, size=10, bold=True, color=navy)
            elif "率" in metric_name:
                trend_sheet.cell(comparison_row, 1).alignment = Alignment(
                    horizontal="left", vertical="center", indent=1
                )
            favorability = _cell_text(record_value(item, "favorability", "favorable_unfavorable"))
            if favorability.upper() in {"F", "FAVORABLE", "有利"}:
                trend_sheet.cell(comparison_row, 7).fill = PatternFill("solid", fgColor=pass_fill)
                trend_sheet.cell(comparison_row, 7).font = Font(
                    name=font_name, size=10, bold=True, color=pass_text
                )
            elif favorability.upper() in {"U", "UNFAVORABLE", "不利"}:
                trend_sheet.cell(comparison_row, 7).fill = PatternFill("solid", fgColor=fail_fill)
                trend_sheet.cell(comparison_row, 7).font = Font(
                    name=font_name, size=10, bold=True, color=fail_text
                )
        if not comparison_rows:
            trend_sheet.cell(4, 1, "未提供可靠的对比数据")

        trend_start = comparison_end + 3
        trend_rows = [
            [
                analysis_cell_value(record_value(item, "period", "date", "label"), "未提供周期"),
                metric_label_with_unit(item),
                analysis_cell_value(record_value(item, "current_value", "value")),
                analysis_cell_value(record_value(item, "baseline_value"), "暂无可比数据"),
                analysis_cell_value(record_value(item, "absolute_change"), "不适用"),
                analysis_cell_value(record_value(item, "relative_change"), "不适用"),
                analysis_cell_value(record_value(item, "anomaly", "note"), "无异常说明"),
            ]
            for item in trends
        ]
        trend_end = write_table(
            trend_sheet,
            start_row=trend_start,
            headers=["周期", "指标", "当前值", "基准值", "绝对变化", "变化率", "异常说明"],
            records=trend_rows,
            widths=[18, 22, 16, 16, 16, 14, 32],
            percent_columns={6},
        )
        for row_offset, item in enumerate(trends, start=1):
            metric_name = _cell_text(record_value(item, "metric", "name"))
            unit = _cell_text(record_value(item, "unit"))
            for column_index in (3, 4, 5):
                value_cell = trend_sheet.cell(trend_start + row_offset, column_index)
                if isinstance(value_cell.value, (int, float)):
                    apply_metric_number_format(value_cell, metric_name, unit)
        if not trend_rows:
            trend_sheet.cell(trend_start + 1, 1, "未提供可靠的趋势数据")
        trend_sheet.freeze_panes = "A4"
        trend_sheet.auto_filter.ref = f"A{trend_start}:G{max(trend_start, trend_end)}"

        # Every comparison chart belongs below the complete visible data area.
        # Keep this anchor independent from helper ranges and from whether the
        # trend table contains data or only its explicit no-data message.
        visible_data_end = max(
            comparison_end,
            trend_end,
            trend_start + 1 if not trend_rows else trend_end,
        )
        chart_anchor_row = visible_data_end + 2

        if trend_rows:
            first_metric = trend_rows[0][1]
            chart_rows = [
                row
                for row in trend_rows
                if row[1] == first_metric and isinstance(row[2], (int, float))
            ][:20]
            period_ordinals: list[int] = []
            for row in chart_rows:
                period_match = re.search(
                    r"(\d{4})\D?(\d{1,2})\D?(\d{1,2})",
                    _cell_text(row[0]).strip(),
                )
                if period_match is None:
                    period_ordinals = []
                    break
                try:
                    period_ordinals.append(
                        date(*(int(part) for part in period_match.groups())).toordinal()
                    )
                except ValueError:
                    period_ordinals = []
                    break
            if len(period_ordinals) == len(chart_rows):
                chart_rows = [
                    row
                    for _, row in sorted(
                        zip(period_ordinals, chart_rows, strict=True),
                        key=lambda item: item[0],
                    )
                ]
            if len(chart_rows) >= 2:
                chart_start = trend_start
                helper_column = 9
                chart_unit = _cell_text(record_value(trends[0], "unit"))
                chart_is_rate = "率" in _cell_text(first_metric) or chart_unit in {
                    "%",
                    "百分比",
                    "百分点",
                }
                trend_sheet.column_dimensions[get_column_letter(helper_column)].width = 16
                trend_sheet.column_dimensions[get_column_letter(helper_column + 1)].width = 18
                for index, row in enumerate(chart_rows, start=1):
                    trend_sheet.cell(chart_start + index, helper_column, row[0])
                    chart_value = row[2]
                    if chart_is_rate and isinstance(chart_value, (int, float)):
                        chart_value = float(chart_value) * 100
                    value_cell = trend_sheet.cell(
                        chart_start + index,
                        helper_column + 1,
                        chart_value,
                    )
                    if chart_is_rate:
                        value_cell.number_format = '0.0"%"'
                trend_sheet.column_dimensions[get_column_letter(helper_column)].hidden = True
                trend_sheet.column_dimensions[get_column_letter(helper_column + 1)].hidden = True
                line_chart = LineChart()
                line_chart.title = f"{_cell_text(first_metric)}趋势"
                line_chart.height = 8
                line_chart.width = 15
                line_chart.legend = None
                line_chart.visible_cells_only = False
                line_chart.y_axis.majorGridlines = None
                line_chart.y_axis.title = chart_unit
                if chart_is_rate:
                    line_chart.y_axis.numFmt = '0"%"'
                line_chart.x_axis.title = "周期"
                line_chart.add_data(
                    Reference(
                        trend_sheet,
                        min_col=helper_column + 1,
                        min_row=chart_start + 1,
                        max_row=chart_start + len(chart_rows),
                    ),
                    titles_from_data=False,
                )
                line_chart.series[0].graphicalProperties.line.solidFill = core_blue
                line_chart.series[0].graphicalProperties.line.width = 24000
                line_chart.set_categories(
                    Reference(
                        trend_sheet,
                        min_col=helper_column,
                        min_row=chart_start + 1,
                        max_row=chart_start + len(chart_rows),
                    )
                )
                # Anchor below every visible table row, not below filtered helper data.
                trend_sheet.add_chart(line_chart, f"A{chart_anchor_row}")
        elif comparison_rows:
            chart = BarChart()
            chart.type = "col"
            chart.title = "当前值与基准值"
            chart.height = 8
            chart.width = 15
            chart.y_axis.majorGridlines = None
            chart.add_data(
                Reference(
                    trend_sheet, min_col=2, max_col=3, min_row=3, max_row=3 + len(comparison_rows)
                ),
                titles_from_data=True,
            )
            for series, color in zip(chart.series, (core_blue, chart_secondary), strict=False):
                series.graphicalProperties.solidFill = color
                series.graphicalProperties.line.solidFill = color
            chart.set_categories(
                Reference(trend_sheet, min_col=1, min_row=4, max_row=3 + len(comparison_rows))
            )
            # Comparison-only reports follow the same layout rule: charts start
            # at column A below the tables, never beside them in columns I+.
            trend_sheet.add_chart(chart, f"A{chart_anchor_row}")

    if analysis_type == "diagnostic" and (anomalies or drivers or recommendations):
        driver_sheet = workbook.create_sheet("异常与建议")
        configure_sheet(driver_sheet)
        add_sheet_title(driver_sheet, "异常与建议", 3)

        issue_rows: list[list[Any]] = []
        for item in anomalies:
            statement = record_value(item, "statement", "anomaly", "name")
            if statement is None:
                continue
            metric_refs = item.get("metric_refs") if isinstance(item.get("metric_refs"), list) else []
            metric_text = joined_unique(
                [_analysis_metric_label(value) for value in metric_refs]
            ) or "相关经营指标"
            observed = record_value(item, "observed_value", "current_value", "value")
            baseline = record_value(item, "baseline_value")
            severity = _xlsx_display_value(record_value(item, "severity"))
            key_data = joined_unique(
                [
                    f"当前值：{_cell_text(observed)}" if observed is not None else None,
                    f"基准值：{_cell_text(baseline)}" if baseline is not None else None,
                    f"严重程度：{_cell_text(severity)}" if severity is not None else None,
                ]
            ) or "需结合业务情况进一步确认"
            issue_rows.append([statement, metric_text, key_data])

        chart_value_candidates: list[tuple[str, float]] = []
        chart_rate_candidates: list[tuple[str, float]] = []
        for item in drivers:
            statement = record_value(item, "statement")
            evidence = meaningful_value(record_value(item, "evidence"))
            factor = record_value(item, "member", "name", "dimension") or statement
            if factor is None:
                continue
            metric_refs = item.get("metric_refs") if isinstance(item.get("metric_refs"), list) else []
            metric_text = record_value(item, "metric") or joined_unique(
                [_analysis_metric_label(value) for value in metric_refs]
            ) or "相关经营指标"
            contribution_value = meaningful_value(record_value(item, "contribution_value"))
            contribution_rate = meaningful_value(
                record_value(item, "contribution_rate", "contribution")
            )
            current_value = meaningful_value(record_value(item, "current_value", "value"))
            key_data = joined_unique(
                [
                    (
                        f"贡献值：{format_metric_display(contribution_value, '', metric_text)}"
                        if contribution_value is not None
                        else None
                    ),
                    (
                        f"贡献占比：{format_metric_display(contribution_rate, '%', '贡献占比')}"
                        if contribution_rate is not None
                        else None
                    ),
                    f"当前值：{_cell_text(current_value)}" if current_value is not None else None,
                    evidence,
                    statement if statement != factor else None,
                ]
            )
            if key_data is None:
                continue
            issue_rows.append([factor, metric_text, key_data])
            if isinstance(contribution_value, (int, float)):
                chart_value_candidates.append((_cell_text(factor), float(contribution_value)))
            if isinstance(contribution_rate, (int, float)):
                chart_rate_candidates.append((_cell_text(factor), float(contribution_rate)))

        issue_end = write_table(
            driver_sheet,
            start_row=3,
            headers=["异常/影响因素", "相关指标", "关键数据"],
            records=issue_rows,
            widths=[30, 24, 58],
        )
        driver_sheet.freeze_panes = "A4"
        driver_sheet.auto_filter.ref = f"A3:C{max(3, issue_end)}"
        if not issue_rows:
            driver_sheet.cell(4, 1, "暂无有数据证据支持的异常或影响因素")

        recommendation_start = issue_end + 3
        recommendation_rows = [
            [
                analysis_cell_value(
                    _xlsx_display_value(record_value(item, "priority")), "中"
                ),
                analysis_cell_value(record_value(item, "action", "statement"), "待补充建议"),
                analysis_cell_value(
                    record_value(item, "rationale", "evidence", "basis"),
                    "依据本报告数据",
                ),
            ]
            for item in recommendations
            if record_value(item, "action", "statement") is not None
        ]
        recommendation_end = write_table(
            driver_sheet,
            start_row=recommendation_start,
            headers=["优先级", "建议动作", "数据依据"],
            records=recommendation_rows,
            widths=[12, 42, 52],
        )
        if not recommendation_rows:
            driver_sheet.cell(recommendation_start + 1, 1, "暂无有数据依据的建议动作")
        driver_sheet.column_dimensions["A"].width = 30
        driver_sheet.column_dimensions["B"].width = 32
        driver_sheet.column_dimensions["C"].width = 58

        chartable_rows = (
            chart_value_candidates[:10]
            if len(chart_value_candidates) >= 2
            else chart_rate_candidates[:10]
        )
        chart_is_rate = len(chart_value_candidates) < 2 and len(chart_rate_candidates) >= 2
        if len(chartable_rows) >= 2:
            helper_label_column = 11
            helper_value_column = 12
            helper_header_row = 3
            helper_first_row = helper_header_row + 1
            for helper_offset, (label, value) in enumerate(chartable_rows, start=1):
                driver_sheet.cell(
                    helper_header_row + helper_offset,
                    helper_label_column,
                    label,
                )
                chart_value = value * 100 if chart_is_rate else value
                value_cell = driver_sheet.cell(
                    helper_header_row + helper_offset,
                    helper_value_column,
                    chart_value,
                )
                if chart_is_rate:
                    value_cell.number_format = '0.0"%"'
            driver_sheet.column_dimensions[
                get_column_letter(helper_label_column)
            ].hidden = True
            driver_sheet.column_dimensions[
                get_column_letter(helper_value_column)
            ].hidden = True
            driver_chart = BarChart()
            driver_chart.type = "bar"
            driver_chart.title = "主要贡献因素"
            driver_chart.height = 8
            driver_chart.width = 15
            driver_chart.legend = None
            driver_chart.visible_cells_only = False
            driver_chart.x_axis.majorGridlines = None
            if chart_is_rate:
                driver_chart.x_axis.numFmt = '0"%"'
            driver_chart.add_data(
                Reference(
                    driver_sheet,
                    min_col=helper_value_column,
                    min_row=helper_first_row,
                    max_row=helper_header_row + len(chartable_rows),
                ),
                titles_from_data=False,
            )
            driver_chart.series[0].graphicalProperties.solidFill = core_blue
            driver_chart.series[0].graphicalProperties.line.solidFill = core_blue
            driver_chart.set_categories(
                Reference(
                    driver_sheet,
                    min_col=helper_label_column,
                    min_row=helper_first_row,
                    max_row=helper_header_row + len(chartable_rows),
                )
            )
            chart_anchor_row = max(recommendation_end, issue_end) + 3
            driver_sheet.add_chart(driver_chart, f"A{chart_anchor_row}")
            print_end_row = chart_anchor_row + 15
        else:
            print_end_row = max(recommendation_end, issue_end, 18)
        driver_sheet.print_area = f"A1:C{print_end_row}"

    detail_sheet = workbook.create_sheet("数据明细")
    detail_sheet.sheet_properties.tabColor = teal
    detail_sheet.sheet_view.showGridLines = False
    detail_sheet.sheet_view.zoomScale = 85
    detail_sheet.sheet_format.defaultRowHeight = 22
    detail_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    detail_sheet.page_setup.orientation = "landscape"
    detail_sheet.page_setup.paperSize = detail_sheet.PAPERSIZE_A4
    detail_sheet.page_setup.fitToWidth = 1
    detail_sheet.page_setup.fitToHeight = 0
    detail_sheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2,
    )

    if columns:
        detail_columns = _merchant_detail_columns(columns)
        display_columns = _display_columns(detail_columns)
        number_kinds = _column_number_kinds(detail_columns)
        column_widths = _column_widths(detail_columns, rows)
        for column_index, label in enumerate(display_columns, start=1):
            cell = detail_sheet.cell(
                row=1,
                column=column_index,
                value=_xlsx_safe_value(
                    label,
                    context=f"Column header {column_index}",
                ),
            )
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor=teal)
            cell.border = cell_border
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )
        detail_sheet.row_dimensions[1].height = 38

        number_formats = {
            "integer": "#,##0;[Red](#,##0);-",
            "decimal": "#,##0.00;[Red](#,##0.00);-",
            "percent": "0.00%;[Red](0.00%);-",
            "percent_point": '0.00"%";[Red](0.00"%");-',
            "currency": "¥#,##0.00;[Red](¥#,##0.00);-",
        }
        for row_index, row in enumerate(rows, start=2):
            row_fill = PatternFill("solid", fgColor=banded if row_index % 2 else white)
            max_wrapped_lines = 1
            for column_index, column in enumerate(detail_columns, start=1):
                source_value = row.get(column)
                display_value = _xlsx_detail_display_value(column, source_value)
                identifier_column = _is_identifier_column(column)
                safe_value = _xlsx_safe_value(
                    display_value,
                    context=f"Data row {row_index - 1}, column {column}",
                    force_text=identifier_column,
                )
                cell = detail_sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=safe_value,
                )
                kind = number_kinds[column_index - 1]
                cell.font = body_font
                cell.fill = row_fill
                cell.border = cell_border
                if isinstance(safe_value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
                elif isinstance(safe_value, date):
                    cell.number_format = "yyyy-mm-dd"
                elif (
                    kind
                    and isinstance(safe_value, (int, float))
                    and not isinstance(safe_value, bool)
                ):
                    cell.number_format = number_formats[kind]
                elif identifier_column:
                    cell.number_format = "@"
                elif isinstance(source_value, (Decimal, int)) and isinstance(safe_value, str):
                    # Excel can only preserve 15 significant numeric digits. Keep
                    # higher-precision values as text instead of silently rounding.
                    cell.number_format = "@"
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )
                if safe_value not in (None, ""):
                    max_wrapped_lines = max(
                        max_wrapped_lines,
                        _wrapped_line_count(
                            safe_value,
                            column_widths[column_index - 1],
                        ),
                    )
            detail_sheet.row_dimensions[row_index].height = float(
                min(300, max(24, 16 * max_wrapped_lines + 8))
            )

        last_column = get_column_letter(len(detail_columns))
        last_row = max(1, len(rows) + 1)
        detail_sheet.freeze_panes = "D2" if len(detail_columns) >= 3 else "B2"
        detail_sheet.auto_filter.ref = f"A1:{last_column}{last_row}"
        if rows and len(set(display_columns)) == len(display_columns):
            raw_table = Table(displayName="RawQueryResult", ref=f"A1:{last_column}{last_row}")
            raw_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            detail_sheet.add_table(raw_table)
        status_fills = {
            "通过": (pass_fill, pass_text),
            "警告": (warning_fill, warning_text),
            "失败": (fail_fill, fail_text),
        }
        for column_index, column in enumerate(detail_columns, start=1):
            normalized = str(column).strip("`\"'").lower()
            if not any(token in normalized for token in ("status", "state", "severity")):
                continue
            column_letter = get_column_letter(column_index)
            target_range = f"{column_letter}2:{column_letter}{last_row}"
            for status, (fill_color, font_color) in status_fills.items():
                detail_sheet.conditional_formatting.add(
                    target_range,
                    FormulaRule(
                        formula=[f'ISNUMBER(SEARCH("{status}",{column_letter}2))'],
                        fill=PatternFill("solid", fgColor=fill_color),
                        font=Font(name=font_name, size=10, bold=True, color=font_color),
                    ),
                )
        detail_sheet.print_title_rows = "1:1"
        detail_sheet.print_area = f"A1:{last_column}{last_row}"
        for column_index, width in enumerate(column_widths, start=1):
            detail_sheet.column_dimensions[get_column_letter(column_index)].width = width
    else:
        empty_cell = detail_sheet["A1"]
        empty_cell.value = "查询结果为空"
        empty_cell.font = label_font
        empty_cell.fill = PatternFill("solid", fgColor=light_teal)
        empty_cell.border = cell_border
        empty_cell.alignment = Alignment(horizontal="left", vertical="center")
        detail_sheet.column_dimensions["A"].width = 24
        detail_sheet.row_dimensions[1].height = 30

    detail_sheet.oddFooter.center.text = "直播经营数据分析报告"
    detail_sheet.oddFooter.right.text = "第 &[Page] 页"
    detail_sheet.oddFooter.center.size = 8
    detail_sheet.oddFooter.center.color = muted
    detail_sheet.oddFooter.right.size = 8
    detail_sheet.oddFooter.right.color = muted

    quality_sheet = workbook.create_sheet("口径与提示")
    configure_sheet(quality_sheet)
    add_sheet_title(quality_sheet, "口径与提示", 6)

    period_coverage = (
        data_quality.get("period_coverage")
        if isinstance(data_quality.get("period_coverage"), dict)
        else {}
    )
    scope_rows: list[list[Any]] = []

    def add_scope_row(label: str, value: Any) -> None:
        if value in (None, "", [], {}):
            return
        scope_rows.append([label, display_analysis_value(value)])

    add_scope_row("统计周期", report_period if report_period != "未提供" else None)
    observed_period = record_value(data_scope, "observed_period", "actual_period")
    if observed_period and _cell_text(observed_period) != _cell_text(report_period):
        add_scope_row("实际覆盖周期", observed_period)
    add_scope_row("数据更新至", data_as_of if data_as_of != "未提供" else None)
    add_scope_row("数据粒度", record_value(data_scope, "grain", "granularity"))
    add_scope_row("直播场次数", record_value(data_scope, "session_count", "sessions"))
    coverage_ratio = record_value(data_scope, "coverage_ratio")
    if coverage_ratio is None:
        coverage_ratio = record_value(period_coverage, "coverage_ratio")
    add_scope_row("周期覆盖率", coverage_ratio)
    add_scope_row("商品明细覆盖范围", data_scope.get("product_table_coverage"))
    filter_values = data_scope.get("filters") if isinstance(data_scope.get("filters"), list) else []
    readable_filters = [text for value in filter_values if (text := merchant_text(value))]
    add_scope_row("关键筛选条件", readable_filters)
    if not scope_rows:
        scope_rows.append(["数据范围", "本次未提供可确认的数据范围"])

    scope_end = write_table(
        quality_sheet,
        start_row=3,
        headers=["数据范围", "内容"],
        records=scope_rows,
        widths=[26, 70],
    )
    for row_offset, (label, _) in enumerate(scope_rows, start=1):
        if label == "周期覆盖率":
            value_cell = quality_sheet.cell(3 + row_offset, 2)
            if isinstance(value_cell.value, (int, float)):
                value_cell.number_format = "0.00%"

    definition_rows = [
        [
            analysis_cell_value(record_value(item, "name"), "未命名指标"),
            metric_definition_text(item),
        ]
        for item in metric_definitions
    ]
    if not definition_rows:
        definition_rows = [["本次查询口径", value] for value in assumptions if merchant_text(value)]
    definition_start = scope_end + 3
    definition_end = write_table(
        quality_sheet,
        start_row=definition_start,
        headers=["指标", "计算口径"],
        records=definition_rows,
        widths=[26, 70],
    )
    if not definition_rows:
        quality_sheet.cell(definition_start + 1, 1, "本次没有需要额外说明的指标口径")

    prompt_rows: list[list[Any]] = []

    def add_prompt_row(kind: str, value: Any) -> None:
        if value in (None, "", [], {}):
            return
        prompt_rows.append([kind, display_analysis_value(value)])

    freshness_details = (
        data_quality.get("freshness") if isinstance(data_quality.get("freshness"), dict) else {}
    )
    lag_days = record_value(freshness_details, "lag_days")
    if isinstance(lag_days, (int, float)) and lag_days > 0:
        add_prompt_row("数据时效", f"数据较当前日期延迟 {lag_days:g} 天")
    if period_coverage.get("complete") is False:
        covered_days = period_coverage.get("covered_days")
        expected_days = period_coverage.get("expected_days")
        coverage_text = (
            f"实际覆盖 {covered_days} 天，应覆盖 {expected_days} 天"
            if covered_days is not None and expected_days is not None
            else "实际数据未完整覆盖请求周期"
        )
        add_prompt_row("周期完整性", coverage_text)
    if truncated or data_quality.get("truncated") is True:
        add_prompt_row("明细完整性", "导出明细达到行数上限，仅包含部分查询结果")
    if data_quality.get("empty_result") is True:
        add_prompt_row("数据结果", "本次查询没有返回可分析的数据")

    null_counts = data_quality.get("null_counts")
    if isinstance(null_counts, dict):
        null_labels = [
            _column_label(_cell_text(key))
            for key, count in null_counts.items()
            if isinstance(count, (int, float)) and count > 0
        ]
        if null_labels:
            add_prompt_row("数据完整性", f"以下字段存在空值：{'、'.join(null_labels)}")
    zero_denominators = data_quality.get("zero_denominators")
    if isinstance(zero_denominators, dict):
        zero_labels = [
            _column_label(_cell_text(key))
            for key, count in zero_denominators.items()
            if isinstance(count, (int, float)) and count > 0
        ]
        if zero_labels:
            add_prompt_row("指标可计算性", f"以下指标存在零分母：{'、'.join(zero_labels)}")
    small_samples = data_quality.get("small_samples")
    if isinstance(small_samples, dict):
        sample_columns = small_samples.get("columns")
        if isinstance(sample_columns, dict):
            sample_labels = [
                _column_label(_cell_text(key))
                for key, count in sample_columns.items()
                if isinstance(count, (int, float)) and count > 0
            ]
            if sample_labels:
                threshold = small_samples.get("threshold")
                threshold_text = f"（阈值 {threshold}）" if threshold is not None else ""
                add_prompt_row(
                    "小样本提示",
                    f"以下维度样本较少{threshold_text}：{'、'.join(sample_labels)}",
                )
    missing_dimensions = data_quality.get("dimensions")
    if isinstance(missing_dimensions, dict):
        missing_values = missing_dimensions.get("missing")
        if isinstance(missing_values, list) and missing_values:
            add_prompt_row(
                "分析限制",
                "缺少以下业务维度："
                + "、".join(_column_label(_cell_text(value)) for value in missing_values),
            )
    for warning in quality_warnings:
        if text := merchant_text(warning):
            add_prompt_row("质量提示", text)
    for note in data_quality.get("notes", []) if isinstance(data_quality.get("notes"), list) else []:
        if text := merchant_text(note):
            add_prompt_row("补充说明", text)
    for note in notes:
        if text := merchant_text(note):
            add_prompt_row("补充说明", text)
    if not prompt_rows:
        prompt_rows.append(["数据状态", "未发现影响本次判断的明显数据问题"])

    prompt_start = definition_end + 3
    prompt_end = write_table(
        quality_sheet,
        start_row=prompt_start,
        headers=["提示类型", "说明"],
        records=prompt_rows,
        widths=[26, 70],
    )
    quality_sheet.freeze_panes = "A4"
    quality_sheet.print_area = f"A1:F{max(prompt_end, prompt_start + 1)}"

    # This controlled exporter intentionally emits no formulas, so LibreOffice
    # recalculation is unnecessary and formula injection remains impossible.
    for worksheet in workbook.worksheets:
        for workbook_row in worksheet.iter_rows():
            for cell in workbook_row:
                if cell.value is not None and cell.alignment.horizontal != "left":
                    alignment = copy(cell.alignment)
                    alignment.horizontal = "left"
                    cell.alignment = alignment
                if getattr(cell, "data_type", None) == "f":
                    raise ValueError(
                        f"Generated XLSX contains a formula at {worksheet.title}!{cell.coordinate}."
                    )

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _json_bytes(
    *,
    title: str,
    summary: str,
    insights: list[str],
    assumptions: list[str],
    notes: list[str],
    sql: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    truncated: bool,
) -> bytes:
    document = {
        "title": title,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "summary": summary,
        "insights": insights,
        "assumptions": assumptions,
        "notes": notes,
        "query": {"sql": sql},
        "data": {
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "truncated": truncated,
        },
    }
    return json.dumps(
        document,
        ensure_ascii=False,
        indent=2,
        default=str,
        allow_nan=False,
    ).encode("utf-8")


def _pdf_paragraph_text(value: Any, *, limit: int = 240) -> str:
    text = _clean_text(_cell_text(value), limit=limit)
    return escape(text).replace("\r\n", "<br/>").replace("\n", "<br/>").replace("\r", "<br/>")


def _pdf_font_path() -> Path:
    configured = os.getenv("XPD_PDF_FONT_PATH", "").strip()
    candidates = [
        Path(configured).expanduser() if configured else None,
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
        Path("/Library/Fonts/Arial Unicode.ttf"),
        Path("/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc"),
        Path("/usr/share/fonts/truetype/arphic/ukai.ttc"),
        Path("C:/Windows/Fonts/msyh.ttc"),
    ]
    for candidate in candidates:
        if candidate is not None and candidate.is_file():
            return candidate.resolve()
    raise RuntimeError(
        "No embeddable Chinese PDF font was found. Configure XPD_PDF_FONT_PATH "
        "with a TrueType Chinese font file."
    )


def _pdf_bytes(
    *,
    title: str,
    summary: str,
    insights: list[str],
    assumptions: list[str],
    notes: list[str],
    sql: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    truncated: bool,
) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (
            LongTable,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("PDF export requires the reportlab package.") from exc

    font_name = "XPDChinese"
    with _export_lock:
        if font_name not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont(font_name, str(_pdf_font_path()), subfontIndex=0))

    output = io.BytesIO()
    page_width, _ = landscape(A4)
    left_margin = right_margin = 14 * mm
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=19 * mm,
        bottomMargin=16 * mm,
        title=title,
        author="直播经营数据分析 Agent",
        subject="经营数据分析报告",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=21,
        leading=29,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#18344A"),
        spaceAfter=8 * mm,
    )
    heading_style = ParagraphStyle(
        "ChineseHeading",
        parent=styles["Heading2"],
        fontName=font_name,
        fontSize=12,
        leading=18,
        textColor=colors.HexColor("#0B7285"),
        spaceBefore=4 * mm,
        spaceAfter=2 * mm,
    )
    body_style = ParagraphStyle(
        "ChineseBody",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=9,
        leading=15,
        textColor=colors.HexColor("#263238"),
        alignment=TA_LEFT,
        wordWrap="CJK",
    )
    small_style = ParagraphStyle(
        "ChineseSmall",
        parent=body_style,
        fontSize=7,
        leading=10,
    )
    table_header_style = ParagraphStyle(
        "ChineseTableHeader",
        parent=small_style,
        fontSize=7,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.white,
    )

    story: list[Any] = [Paragraph(_pdf_paragraph_text(title, limit=200), title_style)]
    metadata = Table(
        [
            ["生成时间", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")],
            ["数据行数", str(len(rows))],
            ["数据完整性", "达到导出上限，仅包含前若干行" if truncated else "未达到导出上限"],
        ],
        colWidths=[28 * mm, 118 * mm],
        hAlign="LEFT",
    )
    metadata.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#52616B")),
                ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor("#18344A")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F7F8")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6E4E5")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D6E4E5")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([metadata, Spacer(1, 3 * mm)])

    def add_text_section(heading: str, values: list[str]) -> None:
        if not values:
            return
        story.append(Paragraph(heading, heading_style))
        for value in values:
            story.append(Paragraph(f"- {_pdf_paragraph_text(value, limit=2000)}", body_style))
            story.append(Spacer(1, 1.2 * mm))

    if summary:
        story.extend(
            [
                Paragraph("经营结论", heading_style),
                Paragraph(_pdf_paragraph_text(summary, limit=MAX_TEXT_CHARS), body_style),
            ]
        )
    add_text_section("关键洞察", insights)
    add_text_section("查询假设", assumptions)
    add_text_section("注意事项", notes)

    story.append(PageBreak())
    story.append(Paragraph("数据明细", heading_style))
    if not columns:
        story.append(Paragraph("（查询结果为空）", body_style))
    else:
        available_width = page_width - left_margin - right_margin
        group_size = 8
        for offset in range(0, len(columns), group_size):
            group = columns[offset : offset + group_size]
            if offset:
                story.append(PageBreak())
                story.append(Paragraph("数据明细（续）", heading_style))
            if len(columns) > group_size:
                story.append(
                    Paragraph(
                        f"字段 {offset + 1}-{offset + len(group)} / {len(columns)}",
                        small_style,
                    )
                )
                story.append(Spacer(1, 1.5 * mm))
            weights = []
            for column in group:
                longest = max(
                    [len(column)] + [len(_cell_text(row.get(column))) for row in rows[:50]]
                )
                weights.append(min(24, max(8, longest)))
            weight_total = sum(weights) or 1
            widths = [available_width * weight / weight_total for weight in weights]
            table_data = [
                [Paragraph(_pdf_paragraph_text(column), table_header_style) for column in group]
            ]
            table_data.extend(
                [Paragraph(_pdf_paragraph_text(row.get(column)), small_style) for column in group]
                for row in rows
            )
            detail_table = LongTable(
                table_data,
                colWidths=widths,
                repeatRows=1,
                hAlign="LEFT",
                splitByRow=1,
            )
            detail_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B7285")),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#C9D7DB")),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#F7FAFA")],
                        ),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            story.append(detail_table)

    story.extend(
        [
            PageBreak(),
            Paragraph("已执行 SQL", heading_style),
            Paragraph(_pdf_paragraph_text(sql, limit=MAX_TEXT_CHARS), small_style),
        ]
    )

    def draw_page(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D6E4E5"))
        canvas.setLineWidth(0.5)
        canvas.line(left_margin, 13 * mm, page_width - right_margin, 13 * mm)
        canvas.setFont(font_name, 7)
        canvas.setFillColor(colors.HexColor("#607D86"))
        canvas.drawString(left_margin, 8.5 * mm, "直播经营数据分析报告")
        canvas.drawRightString(
            page_width - right_margin,
            8.5 * mm,
            f"第 {doc.page} 页",
        )
        canvas.restoreState()

    document.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    return output.getvalue()


def _artifact_bytes(
    output_format: str,
    *,
    title: str,
    summary: str,
    insights: list[str],
    assumptions: list[str],
    notes: list[str],
    sql: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    truncated: bool,
    analysis_type: str = "simple",
    analysis: dict[str, Any] | None = None,
) -> bytes:
    if output_format == "csv":
        return _csv_bytes(columns, rows)
    if output_format == "markdown":
        return _markdown_bytes(
            title=title,
            summary=summary,
            insights=insights,
            assumptions=assumptions,
            notes=notes,
            sql=sql,
            columns=columns,
            rows=rows,
            truncated=truncated,
        )
    if output_format == "pdf":
        return _pdf_bytes(
            title=title,
            summary=summary,
            insights=insights,
            assumptions=assumptions,
            notes=notes,
            sql=sql,
            columns=columns,
            rows=rows,
            truncated=truncated,
        )
    if output_format == "json":
        return _json_bytes(
            title=title,
            summary=summary,
            insights=insights,
            assumptions=assumptions,
            notes=notes,
            sql=sql,
            columns=columns,
            rows=rows,
            truncated=truncated,
        )
    return _xlsx_bytes(
        title=title,
        summary=summary,
        insights=insights,
        assumptions=assumptions,
        notes=notes,
        sql=sql,
        columns=columns,
        rows=rows,
        truncated=truncated,
        analysis_type=analysis_type,
        analysis=analysis,
    )


def _xlsx_values_equal(expected: Any, actual: Any) -> bool:
    if isinstance(expected, float) and isinstance(actual, (int, float)):
        return expected == float(actual)
    return expected == actual


def _validate_artifact_bytes(
    output_format: str,
    content: bytes,
    *,
    title: str = "",
    columns: list[str] | None = None,
    rows: list[dict[str, Any]] | None = None,
    analysis_type: str = "simple",
) -> dict[str, Any]:
    """Validate file structure and, for XLSX, the exported query-cell contents."""
    if output_format == "csv":
        text = content.decode("utf-8-sig")
        for _ in csv.reader(io.StringIO(text, newline="")):
            pass
        return {"status": "passed", "scope": "file_structure"}
    if output_format == "markdown":
        content.decode("utf-8")
        return {"status": "passed", "scope": "file_structure"}
    if output_format == "json":
        parsed = json.loads(content.decode("utf-8"))
        if not isinstance(parsed, dict) or not isinstance(parsed.get("data"), dict):
            raise ValueError("Generated JSON report is incomplete.")
        return {"status": "passed", "scope": "file_structure"}
    if output_format == "pdf":
        try:
            from pypdf import PdfReader
        except ImportError as exc:  # pragma: no cover - deployment dependency guard
            raise RuntimeError("PDF validation requires the pypdf package.") from exc
        reader = PdfReader(io.BytesIO(content))
        if reader.is_encrypted or not reader.pages:
            raise ValueError("Generated PDF report is invalid or empty.")
        return {"status": "passed", "scope": "file_structure"}

    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/styles.xml",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/sheet2.xml",
    }
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        if archive.testzip() is not None or not required <= set(archive.namelist()):
            raise ValueError("Generated XLSX package is incomplete.")
        for name in required:
            if name.endswith(".xml") or name.endswith(".rels"):
                ElementTree.fromstring(archive.read(name))

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("XLSX validation requires the openpyxl package.") from exc

    expected_columns = list(columns or [])
    expected_rows = list(rows or [])
    workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=False)
    try:
        normalized_analysis_type = _clean_analysis_type(analysis_type)
        expected_sheets = ["经营摘要"]
        if (
            normalized_analysis_type in {"comparison", "diagnostic"}
            and "趋势与对比" in workbook.sheetnames
        ):
            expected_sheets.append("趋势与对比")
        if normalized_analysis_type == "diagnostic" and "异常与建议" in workbook.sheetnames:
            expected_sheets.append("异常与建议")
        expected_sheets.extend(["数据明细", "口径与提示"])
        if workbook.sheetnames != expected_sheets:
            raise ValueError("Generated XLSX has unexpected worksheets.")
        summary_sheet = workbook["经营摘要"]
        detail_sheet = workbook["数据明细"]
        expected_title = _xlsx_safe_value(title, context="Report title")
        if summary_sheet["A1"].value != expected_title:
            raise ValueError("Generated XLSX report title does not match the export request.")

        for worksheet in workbook.worksheets:
            for workbook_row in worksheet.iter_rows():
                for cell in workbook_row:
                    if cell.data_type == "f":
                        raise ValueError(
                            f"Generated XLSX contains a formula at "
                            f"{worksheet.title}!{cell.coordinate}."
                        )
                    if isinstance(cell.value, str):
                        if len(cell.value) > EXCEL_MAX_CELL_CHARS:
                            raise ValueError(
                                f"Generated XLSX text exceeds Excel's cell limit at "
                                f"{worksheet.title}!{cell.coordinate}."
                            )
                        if EXCEL_ILLEGAL_CONTROL_PATTERN.search(cell.value):
                            raise ValueError(
                                f"Generated XLSX contains an illegal control character at "
                                f"{worksheet.title}!{cell.coordinate}."
                            )

        def find_header_row(worksheet: Any, header: str) -> int:
            for row_index in range(1, worksheet.max_row + 1):
                if worksheet.cell(row_index, 1).value == header:
                    return row_index
            raise ValueError(
                f"Generated XLSX analysis header is missing in {worksheet.title}: {header}."
            )

        def require_complete_rows(
            worksheet: Any,
            *,
            start_row: int,
            end_row: int,
            end_column: int,
        ) -> None:
            for row_index in range(start_row, end_row + 1):
                first_value = worksheet.cell(row_index, 1).value
                if isinstance(first_value, str) and first_value.startswith(
                    ("未提供", "暂无", "本次没有")
                ):
                    continue
                for column_index in range(1, end_column + 1):
                    cell = worksheet.cell(row_index, column_index)
                    if cell.value in (None, ""):
                        raise ValueError(
                            f"Generated XLSX analysis row is incomplete at "
                            f"{worksheet.title}!{cell.coordinate}."
                        )

        if "趋势与对比" in workbook.sheetnames:
            trend_sheet = workbook["趋势与对比"]
            trend_header_row = find_header_row(trend_sheet, "周期")
            require_complete_rows(
                trend_sheet,
                start_row=4,
                end_row=trend_header_row - 3,
                end_column=7,
            )
            require_complete_rows(
                trend_sheet,
                start_row=trend_header_row + 1,
                end_row=trend_sheet.max_row,
                end_column=7,
            )

        if "异常与建议" in workbook.sheetnames:
            driver_sheet = workbook["异常与建议"]
            recommendation_header_row = find_header_row(driver_sheet, "优先级")
            require_complete_rows(
                driver_sheet,
                start_row=4,
                end_row=recommendation_header_row - 3,
                end_column=3,
            )
            recommendation_end_row = max(
                (
                    row_index
                    for row_index in range(recommendation_header_row + 1, driver_sheet.max_row + 1)
                    if any(
                        driver_sheet.cell(row_index, column_index).value not in (None, "")
                        for column_index in range(1, 4)
                    )
                ),
                default=recommendation_header_row + 1,
            )
            require_complete_rows(
                driver_sheet,
                start_row=recommendation_header_row + 1,
                end_row=recommendation_end_row,
                end_column=3,
            )

        quality_sheet = workbook["口径与提示"]
        definition_header_row = find_header_row(quality_sheet, "指标")
        quality_header_row = find_header_row(quality_sheet, "提示类型")
        require_complete_rows(
            quality_sheet,
            start_row=4,
            end_row=definition_header_row - 3,
            end_column=2,
        )
        require_complete_rows(
            quality_sheet,
            start_row=definition_header_row + 1,
            end_row=quality_header_row - 3,
            end_column=2,
        )
        require_complete_rows(
            quality_sheet,
            start_row=quality_header_row + 1,
            end_row=quality_sheet.max_row,
            end_column=2,
        )

        if expected_columns:
            merchant_columns = _merchant_detail_columns(expected_columns)
            expected_headers = [
                _xlsx_safe_value(label, context=f"Column header {index}")
                for index, label in enumerate(_display_columns(merchant_columns), start=1)
            ]
            actual_headers = [
                detail_sheet.cell(row=1, column=index).value
                for index in range(1, len(merchant_columns) + 1)
            ]
            if actual_headers != expected_headers:
                raise ValueError("Generated XLSX column headers do not match the query result.")
            if detail_sheet.max_row != len(expected_rows) + 1:
                raise ValueError("Generated XLSX row count does not match the query result.")
            for row_index, source_row in enumerate(expected_rows, start=2):
                for column_index, column in enumerate(merchant_columns, start=1):
                    identifier_column = _is_identifier_column(column)
                    expected = _xlsx_safe_value(
                        _xlsx_display_value(source_row.get(column)),
                        context=f"Data row {row_index - 1}, column {column}",
                        force_text=identifier_column,
                    )
                    actual_cell = detail_sheet.cell(row=row_index, column=column_index)
                    if not _xlsx_values_equal(expected, actual_cell.value):
                        raise ValueError(
                            f"Generated XLSX value mismatch at {actual_cell.coordinate}."
                        )
                    if identifier_column and actual_cell.number_format != "@":
                        raise ValueError(
                            f"Generated XLSX identifier is not stored as text at "
                            f"{actual_cell.coordinate}."
                        )
        elif detail_sheet["A1"].value != "查询结果为空":
            raise ValueError("Generated XLSX empty-result marker is missing.")
    finally:
        workbook.close()

    return {
        "status": "passed",
        "scope": "xlsx_structure_query_rows_and_analysis_sheets",
        "analysis_sheets_checked": True,
        "business_conclusions_checked": False,
    }


def _stored_artifact_files(root: Path) -> list[tuple[str, Path]]:
    artifacts: list[tuple[str, Path]] = []
    if not root.exists() or root.is_symlink():
        return artifacts
    for session_dir in root.iterdir():
        if (
            session_dir.is_symlink()
            or not session_dir.is_dir()
            or not SESSION_ID_PATTERN.fullmatch(session_dir.name)
        ):
            continue
        exports_dir = session_dir / "exports"
        if exports_dir.is_symlink() or not exports_dir.is_dir():
            continue
        for path in exports_dir.iterdir():
            if (
                not path.is_symlink()
                and path.is_file()
                and ARTIFACT_FILENAME_PATTERN.fullmatch(path.name)
            ):
                artifacts.append((session_dir.name, path))
    return artifacts


def _cleanup_expired_artifacts(root: Path, *, now: float) -> None:
    retention_days = _bounded_int("XPD_FILE_RETENTION_DAYS", 30, 0)
    if retention_days <= 0:
        return
    cutoff = now - retention_days * 86_400
    for _, path in _stored_artifact_files(root):
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
        except FileNotFoundError:
            continue


def _file_size(path: Path) -> int:
    try:
        return path.stat().st_size
    except FileNotFoundError:
        return 0


def _write_artifact(exports_dir: Path, filename: str, content: bytes) -> tuple[str, Path]:
    max_files = _bounded_int("XPD_FILE_MAX_ARTIFACTS_PER_SESSION", 50, 1)
    max_file_bytes = _bounded_int("XPD_FILE_MAX_BYTES_PER_ARTIFACT", 10 * 1024 * 1024, 1024)
    max_total_bytes = _bounded_int(
        "XPD_FILE_MAX_TOTAL_BYTES_PER_SESSION", 100 * 1024 * 1024, max_file_bytes
    )
    if not content:
        raise ValueError("Refusing to create an empty report artifact.")
    if len(content) > max_file_bytes:
        raise ValueError(f"Report artifact exceeds the {max_file_bytes}-byte limit.")

    storage_root = exports_dir.parent.parent
    with _artifact_storage_lock(storage_root):
        _cleanup_expired_artifacts(storage_root, now=time.time())
        stored = _stored_artifact_files(storage_root)
        existing = [path for session_id, path in stored if session_id == exports_dir.parent.name]
        if len(existing) >= max_files:
            raise ValueError(f"This session already has the maximum of {max_files} report files.")
        total_bytes = sum(_file_size(path) for path in existing)
        if total_bytes + len(content) > max_total_bytes:
            raise ValueError("This session has reached its total report file storage limit.")

        owner_scope = exports_dir.parent.name.split("_", 2)[1]
        owner_limit = _bounded_int(
            "XPD_FILE_MAX_TOTAL_BYTES_PER_OWNER", 500 * 1024 * 1024, max_total_bytes
        )
        root_limit = _bounded_int("XPD_FILE_MAX_TOTAL_BYTES", 5 * 1024 * 1024 * 1024, owner_limit)
        owner_bytes = sum(
            _file_size(path)
            for stored_session_id, path in stored
            if stored_session_id.split("_", 2)[1] == owner_scope
        )
        root_bytes = sum(_file_size(path) for _, path in stored)
        if owner_bytes + len(content) > owner_limit:
            raise ValueError("This owner has reached the total report file storage limit.")
        if root_bytes + len(content) > root_limit:
            raise ValueError("Report file storage has reached its global limit.")
        min_free_bytes = _bounded_int("XPD_FILE_MIN_FREE_BYTES", 256 * 1024 * 1024, 0)
        if shutil.disk_usage(storage_root).free - len(content) < min_free_bytes:
            raise ValueError("Not enough free disk space to create another report file.")

        artifact_id = f"art_{uuid.uuid4().hex}"
        if not ARTIFACT_ID_PATTERN.fullmatch(artifact_id):
            raise RuntimeError("Failed to create a valid report artifact id.")
        target = exports_dir / f"{artifact_id}{FILENAME_SEPARATOR}{filename}"
        temporary = exports_dir / f".{artifact_id}.tmp"
        try:
            with temporary.open("xb") as handle:
                handle.write(content)
                handle.flush()
                os.fsync(handle.fileno())
            os.replace(temporary, target)
            os.chmod(target, 0o600)
        finally:
            temporary.unlink(missing_ok=True)
        return artifact_id, target


def export_report_file(args: dict[str, Any] | None = None, **kwargs: Any) -> str:
    try:
        payload = dict(args or {})
        output_format = str(payload.get("format") or "").strip().lower()
        if output_format not in SUPPORTED_FORMATS:
            raise ValueError("format must be csv, xlsx, markdown, pdf, or json.")
        active_session_id = _session_id(kwargs.get("session_id"), kwargs.get("task_id"))
        title = _clean_text(payload.get("title"), limit=200) or "经营报告"
        result = query_result_registry.get(
            result_id=payload.get("result_id"),
            session_id=active_session_id,
        )
        if result is None:
            raise ValueError(
                "Query result not found, expired, or owned by another session. "
                "Ask the user to run the data query again before exporting."
            )
        sql = str(result["sql"])
        columns = [str(column) for column in result["columns"]]
        rows = list(result["rows"])
        truncated = bool(result["truncated"])
        if len(columns) > MAX_COLUMNS:
            raise ValueError(f"Report export supports at most {MAX_COLUMNS} columns.")
        analysis_type = _clean_analysis_type(
            payload.get("analysis_type"), required=output_format == "xlsx"
        )
        analysis = _clean_analysis(payload.get("analysis"))
        summary = _clean_text(payload.get("summary"))
        insights = _clean_list(payload.get("insights"))
        assumptions = _clean_list(payload.get("assumptions"))
        notes = _clean_list(payload.get("notes"))
        filename = _sanitize_filename(payload.get("filename"), title, output_format)
        content = _artifact_bytes(
            output_format,
            title=title,
            summary=summary,
            insights=insights,
            assumptions=assumptions,
            notes=notes,
            sql=sql,
            columns=columns,
            rows=rows,
            truncated=truncated,
            analysis_type=analysis_type,
            analysis=analysis,
        )
        validation = _validate_artifact_bytes(
            output_format,
            content,
            title=title,
            columns=columns,
            rows=rows,
            analysis_type=analysis_type,
        )
        artifact_id, path = _write_artifact(_exports_dir(active_session_id), filename, content)
        try:
            remote = upload_report_artifact(
                path,
                session_id=active_session_id,
                artifact_id=artifact_id,
                filename=filename,
                media_type=MEDIA_TYPES[output_format],
            )
        except Exception:
            # Export is successful only when the configured durable destination
            # has accepted the object. Do not leave a local-only artifact that
            # the API could mistake for a completed OSS report.
            path.unlink(missing_ok=True)
            raise
        return _json(
            {
                "ok": True,
                "artifact_id": artifact_id,
                "session_id": active_session_id,
                "filename": filename,
                "format": output_format,
                "media_type": MEDIA_TYPES[output_format],
                "size_bytes": len(content),
                "row_count": len(rows),
                "truncated": truncated,
                "result_id": result["result_id"],
                "path": path.name,
                "validated": validation["status"] == "passed",
                "validation": validation,
                **(remote or {}),
                "next_step": (
                    "Return the download_url and file metadata to the user."
                    if remote
                    else "Return the file metadata and tell the user to use the file link."
                ),
            }
        )
    except Exception as exc:
        return _error(exc)
