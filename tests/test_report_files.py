from __future__ import annotations

import csv
import io
import json
import re
import stat
import sys
import types
import zipfile
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from xml.etree import ElementTree

import pytest
from openpyxl import load_workbook

from xpd_report_agent.api.artifact_store import (
    delete_session_artifacts,
    list_session_artifacts,
    resolve_session_artifact,
)
from xpd_report_agent.hermes_plugin.db_query import report_export, report_oss
from xpd_report_agent.hermes_plugin.db_query.query_results import QueryResultRegistry
from xpd_report_agent.runtime import hermes_report_files

SESSION_A = "xpd_0123456789abcdefabcd_session_a"
SESSION_B = "xpd_fedcba9876543210abcd_session_b"


def _decode(payload: str) -> dict:
    return json.loads(payload)


def _registry_with_result() -> tuple[QueryResultRegistry, str]:
    registry = QueryResultRegistry()
    result_id = registry.store(
        session_id=SESSION_A,
        sql="SELECT item_id, pay_amt FROM report ORDER BY pay_amt DESC",
        columns=["item_id", "pay_amt", "=danger", "note"],
        rows=[
            {
                "item_id": "商品-1",
                "pay_amt": Decimal("123.45"),
                "=danger": ' =HYPERLINK("https://invalid")',
                "note": "含,逗号 | 换行\n测试",
            },
            {
                "item_id": "商品-2",
                "pay_amt": None,
                "=danger": "+1+1",
                "note": datetime(2026, 8, 4, 12, 30),
            },
        ],
        truncated=True,
    )
    assert result_id is not None
    return registry, result_id


def _export(tmp_path: Path, monkeypatch, output_format: str) -> tuple[dict, Path]:
    registry, result_id = _registry_with_result()
    monkeypatch.setattr(report_export, "query_result_registry", registry)
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))
    monkeypatch.setenv("XPD_FILE_MIN_FREE_BYTES", "0")
    payload = _decode(
        report_export.export_report_file(
            {
                "format": output_format,
                "filename": "../" + "直播经营分析" * 30 + ".bad",
                "title": "直播经营分析",
                "result_id": result_id,
                "analysis_type": "simple",
                "summary": "成交额领先。",
                "insights": ["商品-1 贡献最高"],
                "assumptions": ["以支付金额计算"],
                "notes": ["导出数据已截断"],
            },
            session_id=SESSION_A,
        )
    )
    assert payload["ok"] is True
    assert payload["validated"] is True
    path = tmp_path / SESSION_A / "exports" / payload["path"]
    assert path.exists()
    assert path.name == payload["path"]
    assert len(payload["filename"].encode("utf-8")) <= report_export.MAX_FILENAME_BYTES
    assert stat.S_IMODE(path.stat().st_mode) == 0o600
    assert stat.S_IMODE(path.parent.stat().st_mode) == 0o700
    return payload, path


def test_csv_export_uses_exact_result_and_neutralizes_formulas(tmp_path, monkeypatch):
    payload, path = _export(tmp_path, monkeypatch, "csv")

    assert payload["row_count"] == 2
    assert payload["truncated"] is True
    assert path.read_bytes().startswith(b"\xef\xbb\xbf")
    rows = list(csv.reader(io.StringIO(path.read_text(encoding="utf-8-sig"))))
    assert rows[0] == ["商品ID", "成交金额", "风险字段", "备注"]
    assert rows[1][0:3] == ["商品-1", "123.45", '\' =HYPERLINK("https://invalid")']
    assert rows[2][1:3] == ["", "'+1+1"]


def test_markdown_export_contains_business_sections_and_escaped_table(tmp_path, monkeypatch):
    payload, path = _export(tmp_path, monkeypatch, "markdown")

    assert payload["filename"].endswith(".md")
    text = path.read_text(encoding="utf-8")
    assert "# 直播经营分析" in text
    assert "## 经营结论" in text
    assert "## 已执行 SQL" in text
    assert "含,逗号 \\| 换行<br>测试" in text
    assert "仅包含前若干行" in text


def test_xlsx_export_is_polished_without_formulas_or_sql_footer(tmp_path, monkeypatch):
    payload, path = _export(tmp_path, monkeypatch, "xlsx")

    assert payload["media_type"].endswith("spreadsheetml.sheet")
    with zipfile.ZipFile(path) as archive:
        assert archive.testzip() is None
        summary = archive.read("xl/worksheets/sheet1.xml")
        detail = archive.read("xl/worksheets/sheet2.xml")
        assert b"<f>" not in detail
        assert b"<f " not in detail
        assert b"SELECT item_id" not in summary
        ElementTree.fromstring(summary)
        ElementTree.fromstring(detail)

    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == ["经营摘要", "数据明细", "数据口径与质量", "查询审计"]
    summary_sheet = workbook["经营摘要"]
    detail_sheet = workbook["数据明细"]
    assert "A1:P1" in {str(cell_range) for cell_range in summary_sheet.merged_cells.ranges}
    assert summary_sheet.sheet_view.showGridLines is False
    assert summary_sheet.freeze_panes == "A4"
    assert summary_sheet["A1"].font.name == "Microsoft YaHei"
    assert summary_sheet["A1"].font.bold is True
    assert summary_sheet["A1"].fill.fgColor.rgb.endswith("1F4E78")
    assert summary_sheet["A3"].alignment.horizontal == "left"
    assert summary_sheet["B2"].alignment.horizontal == "left"
    assert summary_sheet["A2"].value == "报告周期"
    assert summary_sheet["I2"].value == "数据截至"
    assert "模型状态" not in [cell.value for cell in summary_sheet[2]]
    assert "币种" not in [cell.value for cell in summary_sheet[2]]
    assert "数据行数 2" in summary_sheet["A3"].value
    assert detail_sheet.freeze_panes == "D2"
    assert detail_sheet.auto_filter.ref == "A1:D3"
    assert "RawQueryResult" in detail_sheet.tables
    assert detail_sheet["A1"].comment.text == "原始字段：item_id"
    assert detail_sheet.page_setup.orientation == "landscape"
    assert [detail_sheet.cell(1, column).value for column in range(1, 5)] == [
        "商品ID",
        "成交金额",
        "风险字段",
        "备注",
    ]
    assert detail_sheet["B2"].number_format.startswith("¥")
    assert detail_sheet["C2"].value.startswith("'")
    assert detail_sheet["C3"].value.startswith("'")
    assert all(
        cell.data_type != "f"
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )
    workbook.close()


@pytest.mark.parametrize(
    ("analysis_type", "expected_sheets"),
    [
        ("simple", ["经营摘要", "数据明细", "数据口径与质量", "查询审计"]),
        (
            "comparison",
            ["经营摘要", "趋势与对比", "数据明细", "数据口径与质量", "查询审计"],
        ),
        (
            "diagnostic",
            [
                "经营摘要",
                "趋势与对比",
                "驱动分析",
                "数据明细",
                "数据口径与质量",
                "查询审计",
            ],
        ),
    ],
)
def test_xlsx_analysis_type_controls_auditable_sheet_structure(analysis_type, expected_sheets):
    analysis = {
        "metrics": [
            {
                "name": "成交金额",
                "current_value": 120000,
                "baseline_value": 100000,
                "absolute_change": 20000,
                "relative_change": 0.2,
                "unit": "元",
            }
        ],
        "comparisons": [
            {
                "metric": "成交金额",
                "current_value": 120000,
                "baseline_value": 100000,
                "absolute_change": 20000,
                "relative_change": 0.2,
                "unit": "元",
                "baseline_label": "等长上一周期",
                "favorability": "F",
            }
        ],
        "trends": [
            {"period": "2026-08-01", "metric": "成交金额", "value": 50000, "unit": "元"},
            {"period": "2026-08-02", "metric": "成交金额", "value": 70000, "unit": "元"},
        ],
        "drivers": [
            {
                "dimension": "商品",
                "member": "商品A",
                "metric": "成交金额",
                "contribution_value": 15000,
                "contribution_rate": 0.75,
                "statement": "商品A贡献主要增量",
                "evidence": "贡献增量15000元",
                "confidence": 1.0,
            }
        ],
        "recommendations": [
            {"action": "优先补充商品A库存", "evidence": "贡献整体增量75%", "priority": "高"}
        ],
        "metric_definitions": [
            {"name": "成交金额", "formula": "SUM(pay_amt)", "unit": "元", "grain": "自然日"}
        ],
        "data_scope": {
            "period_start": "2026-08-01",
            "period_end": "2026-08-02",
            "source_tables": ["tb_live_goods_daily_stats"],
        },
        "data_quality": {
            "freshness": {"latest": "2026-08-02", "lag_days": 1},
            "period_coverage": {"coverage_ratio": 1.0, "complete": True},
            "warnings": [],
        },
    }
    content = report_export._xlsx_bytes(
        title="动态分析报表",
        summary="成交金额较上一周期增长20%。",
        insights=["商品A贡献主要增量"],
        assumptions=["北京时间自然日"],
        notes=[],
        sql="SELECT stat_date, SUM(pay_amt) FROM tb_live_goods_daily_stats",
        columns=["stat_date", "pay_amt"],
        rows=[{"stat_date": "2026-08-01", "pay_amt": 50000}],
        truncated=False,
        analysis_type=analysis_type,
        analysis=analysis,
    )

    workbook = load_workbook(io.BytesIO(content), data_only=False)
    assert workbook.sheetnames == expected_sheets
    assert workbook["经营摘要"]["A1"].value == "动态分析报表"
    assert workbook["经营摘要"]["A7"].font.name == "Aptos"
    assert workbook["经营摘要"]["A7"].font.size == 20
    assert workbook["经营摘要"]["A6"].font.color.rgb.endswith("1F4E78")
    assert workbook["经营摘要"]["A11"].font.name == "KaiTi"
    assert workbook["经营摘要"]["A11"].font.size == 11
    assert workbook["查询审计"].cell(12, 1).value == "已执行 SQL"
    assert "SUM(pay_amt)" in workbook["查询审计"].cell(13, 1).value
    assert workbook["数据口径与质量"].max_row >= 10
    if analysis_type in {"comparison", "diagnostic"}:
        assert len(workbook["趋势与对比"]._charts) == 1
        assert workbook["趋势与对比"]["H3"].value == "有利/不利"
        assert workbook["趋势与对比"]["H4"].value == "有利"
        assert workbook["趋势与对比"]["H4"].fill.fgColor.rgb.endswith("E2F0D9")
    if analysis_type == "diagnostic":
        assert len(workbook["驱动分析"]._charts) == 1
        assert "主要贡献因素" in str(workbook["驱动分析"]._charts[0].title)
    assert all(
        cell.data_type != "f"
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )
    workbook.close()


def test_xlsx_export_rejects_missing_analysis_type(tmp_path, monkeypatch):
    registry, result_id = _registry_with_result()
    monkeypatch.setattr(report_export, "query_result_registry", registry)
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))

    payload = _decode(
        report_export.export_report_file(
            {"format": "xlsx", "title": "未确认类型", "result_id": result_id},
            session_id=SESSION_A,
        )
    )

    assert payload["ok"] is False
    assert "analysis_type must be simple, comparison, or diagnostic" in payload["error"]


def test_xlsx_normalizes_legacy_analysis_without_empty_cells_or_english_labels():
    analysis = {
        "comparisons": [
            {},
            {"dimension": "整体", "退货件数": 4396, "件数退货率": 0.1294},
        ],
        "trends": [
            {},
            {
                "date": "2026-07-29",
                "session": "鞋靴场",
                "退货件数": 100,
                "件数退货率": 0.1455,
            },
            {
                "date": "2026-07-30",
                "session": "香水场",
                "退货件数": 0,
                "件数退货率": 0.12,
            },
        ],
        "drivers": [
            {},
            {
                "driver": "高客单商品拉高金额退货率",
                "evidence": "金额退货率19.73%",
                "impact": "高于件数退货率6.8个百分点",
            },
        ],
        "metric_definitions": [
            {},
            {
                "metric": "件数退货率",
                "formula": "SUM(refund_itm_qty)/SUM(pay_itm_qty)",
            },
        ],
        "data_scope": {
            "requested_period": "最近30天",
            "observed_period": "2026-07-01至2026-07-30",
            "session_count": 12,
            "granularity": "直播场次",
        },
        "data_quality": {
            "freshness": "PASS",
            "coverage": "完整",
            "small_sample": "无",
        },
    }
    columns = ["start_date", "refund_qty_rate", "refund_amt_rate"]
    rows = [
        {
            "start_date": "2026-07-29",
            "refund_qty_rate": 0.1455,
            "refund_amt_rate": 0.1973,
        }
    ]
    content = report_export._xlsx_bytes(
        title="最近30天直播场次退货诊断分析",
        summary="退货表现存在场次差异。",
        insights=[],
        assumptions=[],
        notes=[],
        sql="SELECT start_date, refund_qty_rate, refund_amt_rate FROM report",
        columns=columns,
        rows=rows,
        truncated=False,
        analysis_type="diagnostic",
        analysis=analysis,
    )

    validation = report_export._validate_artifact_bytes(
        "xlsx",
        content,
        title="最近30天直播场次退货诊断分析",
        columns=columns,
        rows=rows,
        analysis_type="diagnostic",
    )
    assert validation["analysis_sheets_checked"] is True

    workbook = load_workbook(io.BytesIO(content), data_only=False)
    trend_sheet = workbook["趋势与对比"]
    trend_header = next(
        row for row in range(1, trend_sheet.max_row + 1) if trend_sheet.cell(row, 1).value == "周期"
    )
    comparison_rows = list(trend_sheet.iter_rows(min_row=4, max_row=trend_header - 3, max_col=8))
    assert len(comparison_rows) == 2
    assert {row[0].value for row in comparison_rows} == {
        "整体｜退货件数",
        "整体｜件数退货率",
    }
    assert all(cell.value not in (None, "") for row in comparison_rows for cell in row)

    trend_rows = list(
        trend_sheet.iter_rows(min_row=trend_header + 1, max_row=trend_sheet.max_row, max_col=7)
    )
    assert len(trend_rows) == 4
    assert any(row[2].value == 0 for row in trend_rows)
    assert all(cell.value not in (None, "") for row in trend_rows for cell in row)

    driver_sheet = workbook["驱动分析"]
    assert driver_sheet["A4"].value == "高于件数退货率6.8个百分点"
    assert driver_sheet["G4"].value == "高客单商品拉高金额退货率"
    assert driver_sheet["H4"].value == "金额退货率19.73%"
    assert all(driver_sheet.cell(4, column).value not in (None, "") for column in range(1, 10))

    quality_sheet = workbook["数据口径与质量"]
    definition_header = next(
        row
        for row in range(1, quality_sheet.max_row + 1)
        if quality_sheet.cell(row, 1).value == "指标名称"
    )
    assert quality_sheet.cell(definition_header + 1, 1).value == "件数退货率"
    assert all(
        quality_sheet.cell(definition_header + 1, column).value not in (None, "")
        for column in range(1, 8)
    )
    visible_labels = {
        quality_sheet.cell(row, 1).value for row in range(1, quality_sheet.max_row + 1)
    }
    assert {"请求统计周期", "实际覆盖周期", "直播场次数", "数据粒度", "数据时效"} <= visible_labels
    assert (
        not {
            "requested_period",
            "observed_period",
            "session_count",
            "granularity",
            "freshness",
        }
        & visible_labels
    )
    assert "通过" in {
        quality_sheet.cell(row, 2).value for row in range(1, quality_sheet.max_row + 1)
    }

    detail_headers = [workbook["数据明细"].cell(1, column).value for column in range(1, 4)]
    assert detail_headers == ["直播开始日期", "件数退货率", "金额退货率"]
    workbook.close()


def test_xlsx_export_preserves_precision_ids_timezone_and_long_text(tmp_path, monkeypatch):
    registry = QueryResultRegistry()
    columns = [
        "item_id",
        "gmv",
        "exact_decimal",
        "huge_count",
        "live_cnt",
        "total_hours",
        "avg_peak_online",
        "event_time",
        "item_title",
    ]
    long_title = "很长的商品标题" * 20 + "\x01尾部"
    rows = [
        {
            "item_id": 12_345_678_901_234_567_890,
            "gmv": Decimal("105735.40"),
            "exact_decimal": Decimal("999999999999999.99"),
            "huge_count": 12_345_678_901_234_567_890,
            "live_cnt": 1,
            "total_hours": Decimal("18.5"),
            "avg_peak_online": 1338,
            "event_time": datetime(2026, 8, 5, 10, 0, tzinfo=UTC),
            "item_title": long_title,
        }
    ]
    result_id = registry.store(
        session_id=SESSION_A,
        sql="SELECT exact_export_values",
        columns=columns,
        rows=rows,
        truncated=False,
    )
    assert result_id is not None
    monkeypatch.setattr(report_export, "query_result_registry", registry)
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))
    monkeypatch.setenv("XPD_FILE_MIN_FREE_BYTES", "0")
    monkeypatch.setenv("HERMES_TIMEZONE", "Asia/Shanghai")

    payload = _decode(
        report_export.export_report_file(
            {
                "format": "xlsx",
                "title": "边界值报表",
                "result_id": result_id,
                "analysis_type": "simple",
            },
            session_id=SESSION_A,
        )
    )

    assert payload["ok"] is True
    assert payload["validated"] is True
    assert payload["validation"] == {
        "status": "passed",
        "scope": "xlsx_structure_query_rows_and_analysis_sheets",
        "analysis_sheets_checked": True,
        "business_conclusions_checked": False,
    }
    path = tmp_path / SESSION_A / "exports" / payload["path"]
    workbook = load_workbook(path, data_only=False)
    detail = workbook["数据明细"]
    assert [detail.cell(1, index).value for index in range(1, 10)][0:2] == [
        "商品ID",
        "成交金额（GMV）",
    ]
    assert detail["E1"].value == "直播场次数"
    assert detail["F1"].value == "总直播时长（小时）"
    assert detail["G1"].value == "平均峰值在线人数"
    assert detail["A2"].value == "12345678901234567890"
    assert detail["A2"].number_format == "@"
    assert detail["B2"].value == 105735.4
    assert detail["B2"].number_format.startswith("¥")
    assert detail["C2"].value == "999999999999999.99"
    assert detail["C2"].number_format == "@"
    assert detail["D2"].value == "12345678901234567890"
    assert detail["D2"].number_format == "@"
    assert detail["E2"].number_format.startswith("#,##0")
    assert detail["F2"].number_format.startswith("#,##0.00")
    assert detail["G2"].number_format.startswith("#,##0")
    assert detail["H2"].value == datetime(2026, 8, 5, 18, 0)
    assert "\\x01" in detail["I2"].value
    assert detail.row_dimensions[2].height > 24
    workbook.close()


def test_xlsx_export_rejects_overlong_cell_without_silent_truncation(tmp_path, monkeypatch):
    registry = QueryResultRegistry()
    result_id = registry.store(
        session_id=SESSION_A,
        sql="SELECT note",
        columns=["note"],
        rows=[{"note": "x" * (report_export.EXCEL_MAX_CELL_CHARS + 1)}],
        truncated=False,
    )
    assert result_id is not None
    monkeypatch.setattr(report_export, "query_result_registry", registry)
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))
    monkeypatch.setenv("XPD_FILE_MIN_FREE_BYTES", "0")

    payload = _decode(
        report_export.export_report_file(
            {
                "format": "xlsx",
                "title": "超长文本",
                "result_id": result_id,
                "analysis_type": "simple",
            },
            session_id=SESSION_A,
        )
    )

    assert payload["ok"] is False
    assert "32767" in payload["error"]
    assert "not truncated" in payload["error"]
    assert not list(tmp_path.rglob("*.xlsx"))


def test_xlsx_content_validation_detects_changed_query_value():
    columns = ["item_id", "gmv"]
    rows = [{"item_id": "item-1", "gmv": Decimal("12.50")}]
    content = report_export._xlsx_bytes(
        title="校验报表",
        summary="",
        insights=[],
        assumptions=[],
        notes=[],
        sql="SELECT item_id, gmv",
        columns=columns,
        rows=rows,
        truncated=False,
    )
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    workbook["数据明细"]["B2"] = 999
    tampered = io.BytesIO()
    workbook.save(tampered)
    workbook.close()

    with pytest.raises(ValueError, match="value mismatch"):
        report_export._validate_artifact_bytes(
            "xlsx",
            tampered.getvalue(),
            title="校验报表",
            columns=columns,
            rows=rows,
        )


def test_xlsx_content_validation_detects_empty_analysis_cell():
    content = report_export._xlsx_bytes(
        title="分析页校验",
        summary="",
        insights=[],
        assumptions=[],
        notes=[],
        sql="SELECT 1",
        columns=[],
        rows=[],
        truncated=False,
        analysis_type="diagnostic",
        analysis={
            "comparisons": [{"metric": "退货件数", "current_value": 10}],
            "trends": [{"period": "2026-08-01", "metric": "退货件数", "current_value": 10}],
            "drivers": [{"statement": "场次差异", "evidence": "退货件数差异10件"}],
            "metric_definitions": [{"name": "退货件数", "formula": "SUM(refund_itm_qty)"}],
        },
    )
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    workbook["趋势与对比"]["A4"] = None
    tampered = io.BytesIO()
    workbook.save(tampered)
    workbook.close()

    with pytest.raises(ValueError, match="趋势与对比"):
        report_export._validate_artifact_bytes(
            "xlsx",
            tampered.getvalue(),
            title="分析页校验",
            analysis_type="diagnostic",
        )


def test_json_export_contains_structured_metadata_and_exact_rows(tmp_path, monkeypatch):
    payload, path = _export(tmp_path, monkeypatch, "json")

    assert payload["media_type"] == "application/json; charset=utf-8"
    assert payload["filename"].endswith(".json")
    document = json.loads(path.read_text(encoding="utf-8"))
    assert document["title"] == "直播经营分析"
    assert document["query"]["sql"].startswith("SELECT item_id")
    assert document["data"]["columns"] == ["item_id", "pay_amt", "=danger", "note"]
    assert document["data"]["rows"][0]["item_id"] == "商品-1"
    assert document["data"]["rows"][0]["pay_amt"] == "123.45"
    assert document["data"]["row_count"] == 2
    assert document["data"]["truncated"] is True


def test_pdf_export_is_readable_chinese_business_report(tmp_path, monkeypatch):
    from pypdf import PdfReader

    payload, path = _export(tmp_path, monkeypatch, "pdf")

    assert payload["media_type"] == "application/pdf"
    assert payload["filename"].endswith(".pdf")
    assert path.read_bytes().startswith(b"%PDF-")
    reader = PdfReader(path)
    assert len(reader.pages) >= 3
    assert reader.metadata.title == "直播经营分析"
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "直播经营分析" in text
    assert "商品-1" in text
    assert "SELECT item_id" in text


def test_export_rejects_cross_session_result_without_creating_file(tmp_path, monkeypatch):
    registry, result_id = _registry_with_result()
    monkeypatch.setattr(report_export, "query_result_registry", registry)
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))

    result = _decode(
        report_export.export_report_file(
            {"format": "csv", "title": "报表", "result_id": result_id},
            session_id=SESSION_B,
        )
    )

    assert result["ok"] is False
    assert not (tmp_path / SESSION_B).exists()


def test_report_artifact_uploads_to_oss_and_returns_fresh_signed_url(tmp_path, monkeypatch):
    path = tmp_path / "report.csv"
    path.write_text("a,b\n1,2\n", encoding="utf-8")
    monkeypatch.setenv("XPD_REPORT_OSS_ENABLED", "true")
    monkeypatch.setenv("XPD_REPORT_OSS_ENDPOINT", "https://oss-cn-beijing.aliyuncs.com")
    monkeypatch.setenv("XPD_REPORT_OSS_REGION", "cn-beijing")
    monkeypatch.setenv("XPD_REPORT_OSS_BUCKET", "starpartner-biz")
    monkeypatch.setenv("XPD_REPORT_OSS_PREFIX", "public/dev/agent-report-files/")
    monkeypatch.setenv("XPD_REPORT_OSS_ACCESS_KEY_ID", "test-access-key")
    monkeypatch.setenv("XPD_REPORT_OSS_ACCESS_KEY_SECRET", "test-access-secret")
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))

    class FakeRequest:
        def __init__(self, **kwargs):
            self.__dict__.update(kwargs)

    fake_oss = types.SimpleNamespace(
        PutObjectRequest=FakeRequest,
        GetObjectRequest=FakeRequest,
    )

    class FakeClient:
        def __init__(self):
            self.upload = None
            self.presign_count = 0

        def put_object_from_file(self, request, file_path):
            self.upload = (request, file_path)
            return types.SimpleNamespace(status_code=200, etag="etag-1")

        def presign(self, request, **kwargs):
            self.presign_count += 1
            return types.SimpleNamespace(
                url=f"https://signed.example/{request.key}?v={self.presign_count}",
                expiration=datetime.now(UTC) + timedelta(hours=1),
            )

    client = FakeClient()
    monkeypatch.setattr(report_oss, "_client", lambda config: (fake_oss, client))
    artifact_id = "art_" + "c" * 32
    report_oss.write_report_oss_context(
        SESSION_A,
        uid="seller-007",
        trace_id="trace-abc",
    )

    uploaded = report_oss.upload_report_artifact(
        path,
        session_id=SESSION_A,
        artifact_id=artifact_id,
        filename="经营报告.csv",
        media_type="text/csv; charset=utf-8",
    )
    refreshed = report_oss.remote_artifact_payload(
        path,
        session_id=SESSION_A,
        artifact_id=artifact_id,
        filename="经营报告.csv",
    )

    expected_prefix = client.upload[0].key
    assert re.fullmatch(
        r"public/dev/agent-report-files/\d{8}/"
        r"seller-007-trace-abc-\d{10}\.csv",
        expected_prefix,
    )
    assert client.upload[0].bucket == "starpartner-biz"
    assert client.upload[0].key == expected_prefix
    assert uploaded["storage"] == "oss"
    assert uploaded["oss_uri"] == f"oss://starpartner-biz/{expected_prefix}"
    assert uploaded["download_url"].endswith("?v=1")
    assert refreshed["download_url"].endswith("?v=2")


def test_report_oss_object_key_uses_beijing_day_and_second_timestamp(tmp_path, monkeypatch):
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))
    monkeypatch.setenv("HERMES_TIMEZONE", "Asia/Shanghai")
    report_oss.write_report_oss_context(
        SESSION_A,
        uid="uid-1001",
        trace_id="trace-9001",
    )
    config = report_oss.ReportOssConfig(
        enabled=True,
        endpoint="https://oss-cn-beijing.aliyuncs.com",
        region="cn-beijing",
        bucket="starpartner-biz",
        prefix="public/dev/agent-report-files",
        access_key_id="test-key",
        access_key_secret="test-secret",
        security_token=None,
        download_expires_seconds=3600,
    )

    key = report_oss._object_key(
        config,
        session_id=SESSION_A,
        artifact_id="art_" + "d" * 32,
        filename="经营报告.xlsx",
        now=datetime(2026, 8, 5, 16, 1, 2, tzinfo=UTC),
    )

    assert key == ("public/dev/agent-report-files/20260806/uid-1001-trace-9001-1785945662.xlsx")


def test_query_registry_full_read_does_not_evict_and_new_store_evicts_oldest(monkeypatch):
    monkeypatch.setenv("XPD_QUERY_RESULT_MAX_ENTRIES", "10")
    monkeypatch.setenv("XPD_QUERY_RESULT_MAX_PER_SESSION", "20")
    registry = QueryResultRegistry()
    result_ids = [
        registry.store(
            session_id=SESSION_A,
            sql=f"SELECT {index}",
            columns=["value"],
            rows=[{"value": index}],
            truncated=False,
        )
        for index in range(10)
    ]
    assert all(result_ids)
    assert registry.get(result_id=result_ids[0], session_id=SESSION_A) is not None

    newest = registry.store(
        session_id=SESSION_A,
        sql="SELECT 10",
        columns=["value"],
        rows=[{"value": 10}],
        truncated=False,
    )

    assert newest is not None
    assert registry.get(result_id=result_ids[0], session_id=SESSION_A) is None
    assert registry.get(result_id=result_ids[1], session_id=SESSION_A) is not None


def test_artifact_store_rejects_duplicate_ids_and_symlink_delete(tmp_path, monkeypatch):
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))
    exports = tmp_path / SESSION_A / "exports"
    exports.mkdir(parents=True)
    artifact_id = "art_" + "a" * 32
    (exports / f"{artifact_id}__one.csv").write_text("one", encoding="utf-8")
    (exports / f"{artifact_id}__two.csv").write_text("two", encoding="utf-8")
    with pytest.raises(FileNotFoundError):
        resolve_session_artifact(SESSION_A, artifact_id)

    outside = tmp_path / SESSION_B
    outside.mkdir()
    marker = outside / "marker.txt"
    marker.write_text("keep", encoding="utf-8")
    linked_session = "xpd_aaaaaaaaaaaaaaaaaaaa_linked"
    (tmp_path / linked_session).symlink_to(outside, target_is_directory=True)
    with pytest.raises(ValueError):
        delete_session_artifacts(linked_session)
    assert marker.exists()


def test_native_file_guard_only_reads_current_session_generated_basename(tmp_path, monkeypatch):
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))
    exports = tmp_path / SESSION_A / "exports"
    exports.mkdir(parents=True)
    filename = f"art_{'b' * 32}__report.csv"
    report_path = exports / filename
    report_path.write_text("a,b\n1,2\n", encoding="utf-8")
    other = tmp_path / SESSION_B / "exports"
    other.mkdir(parents=True)
    (other / filename).write_text("secret", encoding="utf-8")

    assert hermes_report_files._safe_read_path(filename, SESSION_A) == report_path
    assert (
        hermes_report_files._safe_read_path(f"../{SESSION_B}/exports/{filename}", SESSION_A) is None
    )
    assert hermes_report_files._safe_read_path(str(report_path), SESSION_A) is None
    assert (
        hermes_report_files._owned_session_for_scope(SESSION_A, "0123456789abcdefabcd") == SESSION_A
    )
    assert hermes_report_files._owned_session_for_scope(SESSION_A, "fedcba9876543210abcd") is None


def test_native_registry_handlers_block_writes_and_cross_session_reads(tmp_path, monkeypatch):
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))
    exports = tmp_path / SESSION_A / "exports"
    exports.mkdir(parents=True)
    filename = f"art_{'c' * 32}__report.md"
    (exports / filename).write_text("# report", encoding="utf-8")

    class Entry:
        def __init__(self, handler):
            self.handler = handler

    calls = []

    def original(args, **kwargs):
        calls.append((args, kwargs))
        return json.dumps({"path": args.get("path"), "ok": True})

    entries = {name: Entry(original) for name in hermes_report_files.FILE_TOOL_NAMES}

    class Registry:
        @staticmethod
        def get_entry(name):
            return entries.get(name)

    tools_module = types.ModuleType("tools")
    registry_module = types.ModuleType("tools.registry")
    registry_module.registry = Registry()
    monkeypatch.setitem(sys.modules, "tools", tools_module)
    monkeypatch.setitem(sys.modules, "tools.registry", registry_module)

    hermes_report_files._patch_registry_handlers()

    read_result = _decode(entries["read_file"].handler({"path": filename}, session_id=SESSION_A))
    assert read_result["ok"] is True
    assert Path(read_result["path"]) == exports / filename
    denied_read = _decode(entries["read_file"].handler({"path": filename}, session_id=SESSION_B))
    assert "error" in denied_read
    denied_write = _decode(entries["write_file"].handler({"path": "x"}, session_id=SESSION_A))
    assert "error" in denied_write
    assert len(calls) == 1


def test_agent_tool_restriction_removes_native_mutations_and_unowned_exports():
    def tool(name: str) -> dict:
        return {"type": "function", "function": {"name": name}}

    class Agent:
        def __init__(self):
            self.tools = [
                tool("read_file"),
                tool("write_file"),
                tool("patch"),
                tool("search_files"),
                tool("export_report_file"),
                tool("db_execute_sql"),
            ]
            self.valid_tool_names = {item["function"]["name"] for item in self.tools}

    owned = Agent()
    hermes_report_files._restrict_agent_tools(owned, SESSION_A)
    assert owned.valid_tool_names == {
        "read_file",
        "export_report_file",
        "db_execute_sql",
    }

    unowned = Agent()
    hermes_report_files._restrict_agent_tools(unowned, None)
    assert unowned.valid_tool_names == {"db_execute_sql"}


def test_list_artifacts_ignores_external_symlink(tmp_path, monkeypatch):
    monkeypatch.setenv("XPD_FILE_STORAGE_PATH", str(tmp_path))
    exports = tmp_path / SESSION_A / "exports"
    exports.mkdir(parents=True)
    target = tmp_path / "outside.csv"
    target.write_text("secret", encoding="utf-8")
    (exports / f"art_{'d' * 32}__outside.csv").symlink_to(target)

    assert list_session_artifacts(SESSION_A) == []
